#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/tgps_user/doctor.py
Version: 0.1.1
Updated: 2026-01-11 (SGT)

Purpose
-------
Step 0.6 "doctor" healthcheck for the SellPut user pipeline.

Validates:
A) lcl.user.yml + cloud policy yml load correctly
B) cloud paths exist and ideas Excel is readable (+ required columns exist)
C) ledger exists and schema_version matches expected
D) key tables/views exist
D2) v_actions_latest is queryable (view-signature drift guard)
E) latest ideas snapshot counts look sane + no dupes by idea_key in v_ideas_latest
F) prints newest output files (policy_eval + queue)

Notes
-----
- DuckDB catalog/schema gotcha: we ATTACH ledger as 'ledger' and fully qualify ledger.lcl.*
- idea_snapshots uses source_sha256; v_ideas_latest uses idea_key.
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

EXPECTED_SCHEMA_VERSION = "0.3"

REQUIRED_TABLES_OR_VIEWS = [
    "idea_snapshots",
    "ideas_seen",
    "policy_runs",
    "policy_row_results",
    "actions",
    "broker_order_links",
    "meta_kv",
    "v_ideas_latest",
]


def _ok(msg: str) -> None:
    print(f"✅ {msg}")


def _fail(msg: str) -> None:
    print(f"❌ {msg}")


def _warn(msg: str) -> None:
    print(f"⚠️  {msg}")


def find_repo_root() -> Path:
    """
    Find the TradersGPS repo root by walking up from this file until we see /tgps-user.
    Fallback to $PROJECT_DIR if set.
    """
    env_root = os.environ.get("PROJECT_DIR")
    if env_root:
        p = Path(os.path.expanduser(env_root)).resolve()
        if (p / "tgps-user").is_dir():
            return p

    here = Path(__file__).resolve()
    for p in [here.parent] + list(here.parents):
        if (p / "tgps-user").is_dir():
            return p

    # last resort: cwd
    cwd = Path.cwd().resolve()
    for p in [cwd] + list(cwd.parents):
        if (p / "tgps-user").is_dir():
            return p

    raise FileNotFoundError(
        "Could not locate repo root (folder containing 'tgps-user'). "
        "Run from within the repo, or set PROJECT_DIR."
    )


def load_yaml(path: Path) -> Dict:
    try:
        import yaml  # type: ignore
    except Exception as e:
        raise RuntimeError("PyYAML not installed. Try: pip install pyyaml") from e

    with path.open("r", encoding="utf-8") as f:
        obj = yaml.safe_load(f)
    if not isinstance(obj, dict):
        raise ValueError(f"YAML did not parse to a dict: {path}")
    return obj


def resolve_cloud_path(cloud_root: str, maybe_rel: str) -> Path:
    p = Path(os.path.expanduser(maybe_rel))
    if p.is_absolute():
        return p
    return Path(os.path.expanduser(cloud_root)) / maybe_rel


def try_open_excel(path: Path) -> Tuple[List[str], Optional[str], Optional[List[str]]]:
    """
    Returns: (sheet_names, chosen_sheet, header_row_values)
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl not installed. Try: pip install openpyxl") from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        chosen = "Conservative" if "Conservative" in sheet_names else wb.sheetnames[0]
        ws = wb[chosen]
        header = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            header.append(cell.value)
        header_s = [("" if h is None else str(h).strip()) for h in header]
        return sheet_names, chosen, header_s
    finally:
        wb.close()


def duckdb_attach(ledger_path: Path):
    try:
        import duckdb  # type: ignore
    except Exception as e:
        raise RuntimeError("duckdb not installed in this venv.") from e

    con = duckdb.connect(database=":memory:")
    con.execute(f"ATTACH '{ledger_path.as_posix()}' AS ledger;")
    return con

def list_tables_and_views(con) -> Tuple[set, set]:
    rows = con.execute(
        """
        SELECT table_name, table_type
        FROM information_schema.tables
        WHERE table_catalog = 'ledger'
          AND table_schema = 'lcl'
        """
    ).fetchall()

    tables = {name for (name, typ) in rows if str(typ).upper() == "BASE TABLE"}
    views  = {name for (name, typ) in rows if str(typ).upper() == "VIEW"}
    return tables, views



def newest_file(folder: Path, contains: str) -> Optional[Path]:
    if not folder.is_dir():
        return None
    c = contains.lower()
    candidates: List[Tuple[float, Path]] = []
    for p in folder.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in (".xlsx", ".xlsm", ".xls"):
            continue
        if c not in p.name.lower():
            continue
        try:
            candidates.append((p.stat().st_mtime, p))
        except OSError:
            continue
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def main(argv: List[str]) -> int:
    p = argparse.ArgumentParser(description="TGPS SellPut Step 0.6 doctor / healthcheck")
    p.add_argument(
        "cmd",
        nargs="?",
        default="check",
        choices=["check"],
        help="Only 'check' is supported (prints PASS/FAIL).",
    )
    p.add_argument("--expected-schema", default=EXPECTED_SCHEMA_VERSION)
    args = p.parse_args(argv)

    failures = 0

    # Locate paths
    try:
        repo_root = find_repo_root()
        user_root = repo_root / "tgps-user"
        cfg_user = user_root / "config" / "lcl.user.yml"
        ledger_path = user_root / "ledger" / "lcl.ledger.duckdb"
        out_sellput = user_root / "output" / "sellput"
        _ok(f"Repo root: {repo_root}")
        _ok(f"User root: {user_root}")
    except Exception as e:
        _fail(f"Locate repo root: {e}")
        return 2

    # A) Load lcl.user.yml + policy yml
    user_cfg = {}
    policy_cfg = {}
    policy_path = None
    ideas_xlsx = None

    if not cfg_user.exists():
        _fail(f"A) lcl.user.yml missing: {cfg_user}")
        return 2

    try:
        user_cfg = load_yaml(cfg_user)
        _ok(f"A) Loaded lcl.user.yml: {cfg_user}")
    except Exception as e:
        _fail(f"A) Load lcl.user.yml failed: {e}")
        return 2

    try:
        ideas_source = user_cfg.get("ideas_source", {}) or {}
        policy_source = user_cfg.get("policy_source", {}) or {}
        cloud_root = str(ideas_source.get("cloud_root", "") or "").strip()
        latest_excel = str(ideas_source.get("latest_excel", "") or "").strip()
        global_policy = str(policy_source.get("global_policy", "") or "").strip()

        if not cloud_root or not latest_excel or not global_policy:
            raise ValueError(
                "Missing one of: ideas_source.cloud_root / ideas_source.latest_excel / policy_source.global_policy"
            )

        ideas_xlsx = resolve_cloud_path(cloud_root, latest_excel)
        policy_path = resolve_cloud_path(cloud_root, global_policy)

        _ok(f"A) Resolved cloud ideas Excel: {ideas_xlsx}")
        _ok(f"A) Resolved cloud policy YAML: {policy_path}")
    except Exception as e:
        _fail(f"A) Resolve cloud paths from lcl.user.yml failed: {e}")
        return 2

    if policy_path is None or not policy_path.exists():
        _fail(f"A) Policy YAML missing: {policy_path}")
        failures += 1
    else:
        try:
            policy_cfg = load_yaml(policy_path)
            _ok(f"A) Loaded policy YAML: {policy_path}")
        except Exception as e:
            _fail(f"A) Load policy YAML failed: {e}")
            failures += 1

    # B) Cloud paths exist + Excel readable
    if ideas_xlsx is None or not ideas_xlsx.exists():
        _fail(f"B) Ideas Excel missing: {ideas_xlsx}")
        failures += 1
    else:
        try:
            sheets, chosen, header = try_open_excel(ideas_xlsx)
            _ok(f"B) Ideas Excel readable: {ideas_xlsx.name} (sheets={len(sheets)}, using='{chosen}')")
            if header:
                _ok(f"B) Header columns detected: {len(header)}")
            else:
                _warn("B) Header row appears empty")

            # Required columns check (from policy)
            try:
                req_cols = policy_cfg.get("modules", {}).get("sellput", {}).get("required_columns", [])
                req_cols = [str(x).strip() for x in req_cols if str(x).strip()]
                if req_cols and header:
                    header_set = {h.strip().lower() for h in header if h is not None}
                    missing = [c for c in req_cols if c.strip().lower() not in header_set]
                    if missing:
                        _fail(f"B) Missing required columns in Excel: {missing}")
                        failures += 1
                    else:
                        _ok(f"B) Required columns present ({len(req_cols)})")
                else:
                    _warn("B) Skipped required-columns check (policy or header missing)")
            except Exception as e:
                _warn(f"B) Required-columns check skipped (error): {e}")

        except Exception as e:
            _fail(f"B) Open/read ideas Excel failed: {e}")
            failures += 1

    # C/D/E) Ledger checks
    if not ledger_path.exists():
        _fail(f"C) Ledger DB missing: {ledger_path}")
        failures += 1
    else:
        try:
            con = duckdb_attach(ledger_path)
        except Exception as e:
            _fail(f"C) DuckDB attach failed: {e}")
            return 2

        try:
            # C) schema_version
            try:
                cols = [r[0] for r in con.execute(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_catalog='ledger' AND table_schema='lcl' AND table_name='meta_kv'
                    ORDER BY ordinal_position
                    """
                ).fetchall()]

                key_col = "k" if "k" in cols else ("key" if "key" in cols else None)
                val_col = "v" if "v" in cols else ("value" if "value" in cols else None)
                ts_col = "updated_at" if "updated_at" in cols else None

                if not key_col or not val_col:
                    raise RuntimeError(f"meta_kv columns unexpected: {cols}")

                order_clause = f'ORDER BY "{ts_col}" DESC' if ts_col else ""
                sql = f"""
                    SELECT "{val_col}"
                    FROM ledger.lcl.meta_kv
                    WHERE "{key_col}" = 'schema_version'
                    {order_clause}
                    LIMIT 1
                """
                row = con.execute(sql).fetchone()
                sv = None if row is None else str(row[0]).strip()

                if sv == str(args.expected_schema).strip():
                    _ok(f"C) schema_version OK: {sv}")
                else:
                    _fail(f"C) schema_version mismatch: got={sv!r} expected={args.expected_schema!r}")
                    failures += 1

            except Exception as e:
                _fail(f"C) Read schema_version failed: {e}")
                failures += 1

            # D) key tables/views exist
            try:
                tables, views = list_tables_and_views(con)
                missing = []
                for name in REQUIRED_TABLES_OR_VIEWS:
                    if name not in tables and name not in views:
                        missing.append(name)
                if missing:
                    _fail(f"D) Missing tables/views in ledger.lcl: {missing}")
                    failures += 1
                else:
                    _ok(f"D) Key tables/views present ({len(REQUIRED_TABLES_OR_VIEWS)})")
            except Exception as e:
                _fail(f"D) List tables/views failed: {e}")
                failures += 1

            # D2) v_actions_latest must be queryable (view-signature drift guard)
            try:
                con.execute("SELECT 1 FROM ledger.lcl.v_actions_latest LIMIT 1").fetchone()
                _ok("D2) v_actions_latest query OK")
            except Exception as e:
                _fail(f"D2) v_actions_latest is broken: {e}  | Run: python -m scripts.tgps_user.init_ledger --repair-views")
                failures += 1

            # E) snapshot counts + no dupes by idea_key in v_ideas_latest
            try:
                n_snap = con.execute("SELECT COUNT(*) FROM ledger.lcl.idea_snapshots").fetchone()[0]
                n_latest = con.execute("SELECT COUNT(*) FROM ledger.lcl.v_ideas_latest").fetchone()[0]
                _ok(f"E) Counts: idea_snapshots={n_snap}  v_ideas_latest={n_latest}")
                if n_snap <= 0 or n_latest <= 0:
                    _fail("E) Counts look wrong (one of them is 0). Did sync_ideas run?")
                    failures += 1

                top_sources = con.execute(
                    """
                    SELECT source_sha256, COUNT(*) AS n
                    FROM ledger.lcl.idea_snapshots
                    GROUP BY 1
                    ORDER BY n DESC
                    LIMIT 5
                    """
                ).fetchall()
                _ok(f"E) Top source_sha256 counts (top 5): {top_sources}")

                dupes = con.execute(
                    """
                    SELECT idea_key, COUNT(*) AS c
                    FROM ledger.lcl.v_ideas_latest
                    GROUP BY 1
                    HAVING c > 1
                    """
                ).fetchall()
                if dupes:
                    _fail(f"E) Duplicate idea_key in v_ideas_latest (should be empty): {dupes[:10]}")
                    failures += 1
                else:
                    _ok("E) No duplicate idea_key in v_ideas_latest")
            except Exception as e:
                _fail(f"E) Snapshot/dupe checks failed: {e}")
                failures += 1

        finally:
            try:
                con.close()
            except Exception:
                pass

    # F) Print newest output file paths
    try:
        out_sellput.mkdir(parents=True, exist_ok=True)
        newest_queue = newest_file(out_sellput, "sellput_queue")
        newest_policy = (
            newest_file(out_sellput, "policy_eval")
            or newest_file(out_sellput, "sellput_policy")
            or newest_file(out_sellput, "policy")
        )

        if newest_policy:
            _ok(f"F) Newest policy output: {newest_policy}")
        else:
            _warn(f"F) No policy output found in: {out_sellput}")

        if newest_queue:
            _ok(f"F) Newest queue output:  {newest_queue}")
        else:
            _warn(f"F) No queue output found in: {out_sellput}")

    except Exception as e:
        _warn(f"F) Output discovery failed: {e}")

    if failures == 0:
        print("\n🎉 DOCTOR RESULT: PASS (all checks ok)")
        return 0

    print(f"\n🧯 DOCTOR RESULT: FAIL ({failures} failing check(s))")
    return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
