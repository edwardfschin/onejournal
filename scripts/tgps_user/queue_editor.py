#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/tgps_user/queue_editor.py
Version: 0.2.4
Updated: 2026-01-22 (SGT)

Creates an editable queue from the latest *_queue_sys.xlsx.

Enhancement (0.2.4)
- Adds deterministic broker-agnostic keys:
    base_key, intent_key (qty included), payload_hash (plan-params hash)
- Adds last-known submission info from DuckDB (if available):
    last_status, last_permIds, last_submitted_at
"""

from __future__ import annotations

import argparse
import math
import os
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import duckdb
import pandas as pd
import yaml  # pyyaml

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from common.order_keys import compute_base_key, compute_intent_key, compute_payload_hash

SGT = timezone(timedelta(hours=8))
SHEET = "ALL"

# Required dropdown values (downstream expects these exact strings)
VALID_ACTIONS = ["WATCH", "REVIEW", "SKIP", "EXECUTE"]
VALID_ENTRY_TYPES = ["LIMIT", "MARKET"]
VALID_DURATIONS = ["DAY", "GTC", "GOOD_TILL_CANCEL"]
VALID_ATTACH_EXIT = ["NO", "YES"]
VALID_EXIT_MODES = ["NONE", "TP_ONLY", "STOP_ONLY", "OCO"]
VALID_STOP_TYPES = ["STOP_MARKET", "STOP_LIMIT"]

# User-editable columns to add
EDIT_COLS = [
    "User Action",
    "Qty",
    "Limit Price Override",
    "User Notes",
    "Entry Order Type",
    "Duration",
    "Attach Exit",
    "TP Limit",
    "Exit Mode",
    "Stop Type",
    "Stop Price",
    "Stop Limit Price",
    "Submit",
]

# System columns (computed; not intended for user edit)
SYSTEM_COLS = [
    "base_key",
    "intent_key",
    "payload_hash",
    "last_status",
    "last_permIds",
    "last_submitted_at",
]


# ----------------------------
# Paths + config
# ----------------------------
def _repo_root() -> Path:
    here = Path.cwd().resolve()
    if (here / "tgps-user").exists():
        return here
    for p in [here] + list(here.parents):
        if (p / "tgps-user").exists():
            return p
    raise SystemExit("❌ Could not find repo root (expected 'tgps-user' folder). Run from TradersGPS repo.")


def _cfg_path(user_root: Path) -> Path:
    return user_root / "config" / "lcl.user.yml"


def _ledger_path(user_root: Path) -> Path:
    return user_root / "ledger" / "lcl.ledger.duckdb"


def _outdir_default(user_root: Path, module: str) -> Path:
    # Keep current convention for now (your workflow already points here)
    return user_root / "output" / "sellput"


def _load_cfg(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise SystemExit(f"❌ Missing config: {path}")
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if not isinstance(data, dict):
        raise SystemExit(f"❌ Config must parse to a mapping/dict: {path}")
    return data


def _module(cfg: Dict[str, Any]) -> str:
    m = ((cfg.get("ideas_source") or {}) or {}).get("module", "")
    m = str(m or "").strip()
    return m or "sellput"


def _bot(cfg: Dict[str, Any], module: str) -> Dict[str, Any]:
    # Convention: "<module>_bot" (sellput_bot, sellcall_bot, etc)
    b = cfg.get(f"{module}_bot")
    if isinstance(b, dict):
        return b
    # fallback for today
    b = cfg.get("sellput_bot")
    return b if isinstance(b, dict) else {}


def _latest(glob_pat: str, folder: Path) -> Optional[Path]:
    cands = sorted(folder.glob(glob_pat), key=lambda p: p.stat().st_mtime, reverse=True)
    return cands[0] if cands else None


def _run_id_to_prefix(run_id: str) -> str:
    # YYYYMMDD_HHMMSS -> YYYY-MM-DD_HHMMSS
    rid = (run_id or "").strip()
    if len(rid) >= 15 and "_" in rid:
        d, t = rid.split("_", 1)
        if len(d) == 8 and d.isdigit():
            return f"{d[0:4]}-{d[4:6]}-{d[6:8]}_{t}"
    return rid


def _find_sys_queue(outdir: Path, explicit: str = "") -> Path:
    if explicit:
        p = Path(explicit).expanduser().resolve()
        if not p.exists():
            raise SystemExit(f"❌ sys queue not found: {p}")
        return p

    rid = (os.environ.get("TGPS_RUN_ID") or "").strip()
    if rid:
        prefix = _run_id_to_prefix(rid)
        cands = sorted(outdir.glob(f"*{prefix}*_queue_sys.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if cands:
            return cands[0]

    p = _latest("*_queue_sys.xlsx", outdir)
    if not p:
        raise SystemExit(f"❌ No *_queue_sys.xlsx found in: {outdir}")
    return p


def _derive_edit_path(sysq: Path, outdir: Path) -> Path:
    name = sysq.name
    if name.endswith("_queue_sys.xlsx"):
        name = name.replace("_queue_sys.xlsx", "_queue_edit.xlsx")
    else:
        name = name.replace(".xlsx", "_queue_edit.xlsx")
    return outdir / name


# ----------------------------
# Helpers
# ----------------------------
def _is_blank(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, float) and pd.isna(x):
        return True
    if isinstance(x, str) and not x.strip():
        return True
    return False


def _num(x: Any) -> Optional[float]:
    if _is_blank(x):
        return None
    try:
        return float(x)
    except Exception:
        return None


def _s(x: Any) -> str:
    return "" if _is_blank(x) else str(x).strip()


def _round_down(x: float, tick: float) -> float:
    if tick <= 0:
        return x
    return math.floor((x + 1e-12) / tick) * tick


def _round_up(x: float, tick: float) -> float:
    if tick <= 0:
        return x
    return math.ceil((x - 1e-12) / tick) * tick


def _tick_size(bot: Dict[str, Any]) -> float:
    pricing = bot.get("pricing") if isinstance(bot.get("pricing"), dict) else {}
    orders = bot.get("orders") if isinstance(bot.get("orders"), dict) else {}
    entry = bot.get("entry") if isinstance(bot.get("entry"), dict) else {}
    tick = (
        _num((pricing or {}).get("tick_size"))
        or _num((orders or {}).get("tick_size"))
        or _num((entry or {}).get("rounding_tick"))
        or 0.05
    )
    return float(tick)


def _limit_default_pct_of_ask(bot: Dict[str, Any]) -> float:
    pricing = bot.get("pricing") if isinstance(bot.get("pricing"), dict) else {}
    orders = bot.get("orders") if isinstance(bot.get("orders"), dict) else {}
    entry = bot.get("entry") if isinstance(bot.get("entry"), dict) else {}
    ws = entry.get("wide_spread_rule") if isinstance(entry.get("wide_spread_rule"), dict) else {}

    v = (
        _num((pricing or {}).get("limit_default_pct_of_ask"))
        or _num((orders or {}).get("limit_default_pct_of_ask"))
        or _num((ws or {}).get("lmin_ask_pct"))
        or 0.90
    )
    return float(v)


def _suggest_entry_limit(
    bid: Optional[float],
    ask: Optional[float],
    *,
    bot: Dict[str, Any],
    tick: float,
) -> Optional[float]:
    if bid is None or ask is None or bid <= 0 or ask <= 0:
        return None

    mid = (bid + ask) / 2.0
    abs_spread = ask - bid
    pct_spread = abs_spread / mid if mid > 0 else 0.0

    entry = bot.get("entry") if isinstance(bot.get("entry"), dict) else {}
    ws = entry.get("wide_spread_rule") if isinstance(entry.get("wide_spread_rule"), dict) else {}

    enabled = bool((ws or {}).get("enabled", True))
    pct_thr = float(_num((ws or {}).get("pct_threshold")) or 0.25)
    abs_thr = float(_num((ws or {}).get("abs_threshold")) or 0.25)

    wide = enabled and ((pct_spread >= pct_thr) or (abs_spread >= abs_thr))

    if not wide:
        return _round_up(mid, tick)

    pct = _limit_default_pct_of_ask(bot)
    raw = max(mid, ask * pct)
    return _round_down(raw, tick)


def _exit_from_credit(
    credit: Optional[float],
    *,
    bot: Dict[str, Any],
    tick: float,
    stop_type: str,
) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if credit is None or credit <= 0:
        return None, None, None

    ex = bot.get("exit") if isinstance(bot.get("exit"), dict) else {}
    stop_mult = float(_num((ex or {}).get("stop_mult")) or 3.00)

    tp_factor = float(_num((ex or {}).get("tp_factor")) or 0.80)
    tp = _round_down(credit * tp_factor, tick)

    stop = _round_up(credit * stop_mult, tick)

    stop_type = (stop_type or "").strip().upper()
    if stop_type != "STOP_LIMIT":
        return tp, stop, None

    if "stop_limit_offset_ticks" in (ex or {}):
        off = int(_num((ex or {}).get("stop_limit_offset_ticks")) or 0)
    elif bool((ex or {}).get("stop_limit_equals_stop", False)):
        off = 0
    else:
        off = 1

    stop_limit = _round_up(stop + float(off) * tick, tick)
    return tp, stop, stop_limit


def _derive_right(module: str, df_row: pd.Series) -> str:
    # Prefer column if present; fallback to module naming
    if "Option Type" in df_row.index:
        ot = _s(df_row.get("Option Type")).upper()
        if ot in ("PUT", "P"):
            return "P"
        if ot in ("CALL", "C"):
            return "C"
    ml = (module or "").strip().lower()
    if ml.startswith("sellcall"):
        return "C"
    return "P"


def _expiry_iso(v: Any) -> str:
    ts = pd.to_datetime(v)
    return ts.strftime("%Y-%m-%d")


def _read_df(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path, engine="openpyxl")
    sheet = SHEET if SHEET in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _apply_defaults(df: pd.DataFrame, *, cfg: Dict[str, Any], module: str) -> pd.DataFrame:
    bot = _bot(cfg, module)
    tick = _tick_size(bot)

    ov = cfg.get("overrides") if isinstance(cfg.get("overrides"), dict) else {}
    qty_default = int(_num((ov or {}).get("default_qty")) or 1)

    entry = bot.get("entry") if isinstance(bot.get("entry"), dict) else {}
    entry_type_default = _s((entry or {}).get("order_type") or "LIMIT").upper()
    duration_default = _s((entry or {}).get("duration") or "GOOD_TILL_CANCEL").upper()

    ex = bot.get("exit") if isinstance(bot.get("exit"), dict) else {}
    always_oco = bool((ex or {}).get("always_oco", True))
    attach_default = "YES" if always_oco else "NO"
    exit_mode_default = "OCO" if always_oco else "NONE"
    stop_type_default = _s((ex or {}).get("stop_type") or "STOP_LIMIT").upper()

    module_l = (module or "").strip().lower()

    for c in EDIT_COLS:
        if c not in df.columns:
            df[c] = ""

    for c in SYSTEM_COLS:
        if c not in df.columns:
            df[c] = ""

    for i, r in df.iterrows():
        if _is_blank(r.get("User Action")):
            df.at[i, "User Action"] = "WATCH"
        if _is_blank(r.get("Qty")):
            df.at[i, "Qty"] = qty_default
        if _is_blank(r.get("Entry Order Type")):
            df.at[i, "Entry Order Type"] = entry_type_default
        if _is_blank(r.get("Duration")):
            df.at[i, "Duration"] = duration_default
        if _is_blank(r.get("Attach Exit")):
            df.at[i, "Attach Exit"] = attach_default
        if _is_blank(r.get("Exit Mode")):
            df.at[i, "Exit Mode"] = exit_mode_default
        if _is_blank(r.get("Stop Type")):
            df.at[i, "Stop Type"] = stop_type_default

        bid = _num(r.get("Bid"))
        ask = _num(r.get("Ask"))

        # Default entry credit
        if _is_blank(r.get("Limit Price Override")):
            lim = _suggest_entry_limit(bid, ask, bot=bot, tick=tick)
            if lim is not None:
                df.at[i, "Limit Price Override"] = float(lim)

        attach = _s(df.at[i, "Attach Exit"]).upper()
        mode = _s(df.at[i, "Exit Mode"]).upper()
        stype = _s(df.at[i, "Stop Type"]).upper()

        if attach == "YES" and mode in ("OCO", "TP_ONLY", "STOP_ONLY"):
            credit = _num(df.at[i, "Limit Price Override"])
            tp, stop, stop_lim = _exit_from_credit(credit, bot=bot, tick=tick, stop_type=stype)

            # SELL PUT rule: TP Limit = max(round_down(credit * 0.20), 0.05)
            if module_l == "sellput" and credit is not None and credit > 0:
                tp = max(_round_down(credit * 0.20, tick), 0.05)

            if mode in ("OCO", "TP_ONLY") and _is_blank(r.get("TP Limit")) and tp is not None:
                df.at[i, "TP Limit"] = float(tp)
            if mode in ("OCO", "STOP_ONLY") and _is_blank(r.get("Stop Price")) and stop is not None:
                df.at[i, "Stop Price"] = float(stop)
            if stype == "STOP_LIMIT" and _is_blank(r.get("Stop Limit Price")) and stop_lim is not None:
                df.at[i, "Stop Limit Price"] = float(stop_lim)

    return df


def _compute_keys(df: pd.DataFrame, *, module: str) -> pd.DataFrame:
    """
    Fill base_key / intent_key / payload_hash for every row.
    """
    for i, r in df.iterrows():
        try:
            ticker = _s(r.get("Ticker")).upper()
            expiry = _expiry_iso(r.get("Expiry"))
            right = _derive_right(module, r)
            strike = float(r.get("Strike Found"))
        except Exception:
            # If sys queue row is malformed, leave blank.
            continue

        qty = int(_num(r.get("Qty")) or 1)
        qty = max(qty, 1)

        base_key = compute_base_key(module=module, ticker=ticker, expiry_iso=expiry, right=right, strike=strike)
        intent_key = compute_intent_key(base_key=base_key, qty=qty)

        # Broker-agnostic plan parameters (hashable)
        plan_params = {
            "module": module,
            "ticker": ticker,
            "expiry": expiry,
            "right": right,
            "strike": float(strike),
            "qty": int(qty),
            "entry_order_type": _s(r.get("Entry Order Type")).upper(),
            "entry_limit": _num(r.get("Limit Price Override")),
            "duration": _s(r.get("Duration")).upper(),
            "attach_exit": _s(r.get("Attach Exit")).upper(),
            "exit_mode": _s(r.get("Exit Mode")).upper(),
            "tp_limit": _num(r.get("TP Limit")),
            "stop_type": _s(r.get("Stop Type")).upper(),
            "stop_price": _num(r.get("Stop Price")),
            "stop_limit_price": _num(r.get("Stop Limit Price")),
        }
        ph = compute_payload_hash(plan_params)

        df.at[i, "base_key"] = base_key
        df.at[i, "intent_key"] = intent_key
        df.at[i, "payload_hash"] = ph

    return df


def _backfill_last_status(df: pd.DataFrame, *, user_root: Path, module: str) -> pd.DataFrame:
    """
    Fill last_* columns from ledger.lcl.order_submissions if present.

    Fix:
    DuckDB can treat `lcl` as either a catalog or schema (ambiguous). We attach the
    ledger DB as catalog `ledger`, then query `ledger.lcl.order_submissions`.
    """
    ledger_path = _ledger_path(user_root)
    if not ledger_path.exists():
        return df

    con = duckdb.connect(":memory:")
    try:
        # Prefer READ_ONLY attach if supported; fallback if not.
        try:
            con.execute(f"ATTACH '{ledger_path.as_posix()}' AS ledger (READ_ONLY);")
        except Exception:
            con.execute(f"ATTACH '{ledger_path.as_posix()}' AS ledger;")

        # Ensure table exists (scoped to attached catalog `ledger`)
        tbl = con.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.tables
            WHERE table_catalog='ledger'
              AND table_schema='lcl'
              AND table_name='order_submissions'
            """
        ).fetchone()
        if not tbl or int(tbl[0] or 0) == 0:
            return df

        # Column presence
        cols = con.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_catalog='ledger'
              AND table_schema='lcl'
              AND table_name='order_submissions'
            """
        ).fetchall()
        colset = {c[0] for c in cols}

        has_perm = "perm_ids_json" in colset
        has_base = "base_key" in colset

        if not has_base:
            # Can't match if base_key wasn't stored historically
            return df

        sql = """
        SELECT status,
               {perm},
               submitted_at
        FROM ledger.lcl.order_submissions
        WHERE module = ?
          AND base_key = ?
          AND dry_run = FALSE
        ORDER BY submitted_at DESC
        LIMIT 1
        """.format(
            perm=("perm_ids_json" if has_perm else "CAST(NULL AS VARCHAR)")
        )

        for i, r in df.iterrows():
            bk = _s(r.get("base_key"))
            if not bk:
                continue
            row = con.execute(sql, [module, bk]).fetchone()
            if not row:
                continue
            st, perm, ts = row[0], row[1], row[2]
            df.at[i, "last_status"] = "" if st is None else str(st)
            df.at[i, "last_permIds"] = "" if perm is None else str(perm)
            df.at[i, "last_submitted_at"] = "" if ts is None else str(ts)

        return df

    finally:
        try:
            con.close()
        except Exception:
            pass


def _write_one_sheet(df: pd.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    cols = [str(c) for c in df.columns]
    ws.append(cols)
    for _, r in df.iterrows():
        ws.append([r.get(c) for c in cols])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"

    max_row = len(df) + 1  # incl header

    def add_list_validation(col_name: str, items: List[str]) -> None:
        if col_name not in cols:
            return
        col_letter = get_column_letter(cols.index(col_name) + 1)
        dv = DataValidation(type="list", formula1='"' + ",".join(items) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_row}")

    add_list_validation("User Action", VALID_ACTIONS)
    add_list_validation("Entry Order Type", VALID_ENTRY_TYPES)
    add_list_validation("Duration", VALID_DURATIONS)
    add_list_validation("Attach Exit", VALID_ATTACH_EXIT)
    add_list_validation("Exit Mode", VALID_EXIT_MODES)
    add_list_validation("Stop Type", VALID_STOP_TYPES)

    wb.save(out_path)


# ----------------------------
# CLI
# ----------------------------
def cmd_patch(args: argparse.Namespace) -> int:
    repo = _repo_root()
    user_root = repo / "tgps-user"
    cfg = _load_cfg(_cfg_path(user_root))
    module = args.module or _module(cfg)

    outdir = Path(args.outdir).expanduser().resolve() if args.outdir else _outdir_default(user_root, module)
    outdir.mkdir(parents=True, exist_ok=True)

    sysq = _find_sys_queue(outdir, explicit=args.sys_queue)
    df = _read_df(sysq)

    if "row_hash" not in df.columns:
        raise SystemExit("❌ SYSTEM queue missing 'row_hash' column (ideas_runner.plan should create it).")

    df = _apply_defaults(df, cfg=cfg, module=module)
    df = _compute_keys(df, module=module)

    if not args.no_status:
        df = _backfill_last_status(df, user_root=user_root, module=module)

    out_path = _derive_edit_path(sysq, outdir)
    if out_path.exists():
        stamp = datetime.now(SGT).strftime("%H%M%S")
        out_path = out_path.with_name(out_path.stem + f"_{stamp}" + out_path.suffix)

    _write_one_sheet(df, out_path)
    print(out_path.as_posix())
    return 0


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Queue Editor (writes one-sheet *_queue_edit.xlsx)")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("patch", help="Generate/refresh *_queue_edit.xlsx (prints path only)")
    p.add_argument("--sys-queue", default="", help="Explicit path to *_queue_sys.xlsx (optional)")
    p.add_argument("--outdir", default="", help="Output dir override (optional)")
    p.add_argument("--module", default="", help="Module override (optional)")
    p.add_argument("--no-status", action="store_true", help="Skip ledger lookup for last_* columns")
    p.set_defaults(func=cmd_patch)

    return ap


def main() -> int:
    args = build_parser().parse_args()
    return int(args.func(args) or 0)


if __name__ == "__main__":
    raise SystemExit(main())
