#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/tgps_user/actions_capture.py
Version: 0.3.3 (2026-01-15, SGT)

Step 7 boundary enforcement: validate + import user actions from *_sellput_queue_edit.xlsx

Key changes in 0.3.3
- Limit defaulting updated to match your spec:
    NOT wide -> round_up_to_tick(Mid)
    wide     -> round_down_to_tick(max(Mid, Ask * limit_default_pct_of_ask))
- Wide detection uses BOTH pct_threshold and abs_threshold (fallback-safe)
- Reads Qty from "Qty" (queue_editor) with fallback to legacy "Qty Override"

Hard guarantees
---------------
- VALIDATE FIRST: import refuses to write anything if any EXECUTE row is BLOCKED.
- Apply defaults once (here), store normalized values in ledger actions.
- Step 8 remains last-line safety, not primary validation.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys
import uuid
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import duckdb
import openpyxl
import yaml
import io
import time
from zipfile import BadZipFile


# ---- Repo root bootstrap ----
CODE_DIR = Path(__file__).resolve().parents[2]
if str(CODE_DIR) not in sys.path:
    sys.path.insert(0, str(CODE_DIR))

# Schwab (best-effort quotes)
from client.schwab_admin import AuthClient, TokenStore  # noqa: E402
from client.schwab_api import RestClient, RestSession  # noqa: E402


VALID_ACTIONS = {"WATCH", "REVIEW", "SKIP", "EXECUTE"}
VALID_ENTRY_TYPES = {"LIMIT", "MARKET"}
VALID_DURATIONS = {"DAY", "GTC", "GOOD_TILL_CANCEL"}
VALID_ATTACH_EXIT = {"NO", "YES"}
VALID_EXIT_MODES = {"NONE", "TP_ONLY", "STOP_ONLY", "OCO"}
VALID_STOP_TYPES = {"STOP_MARKET", "STOP_LIMIT"}


# ----------------------------
# Paths / config
# ----------------------------
def _repo_root() -> Path:
    here = Path.cwd().resolve()
    if (here / "tgps-user").exists():
        return here
    for p in [here] + list(here.parents):
        if (p / "tgps-user").exists():
            return p
    raise SystemExit("❌ Could not find repo root (expected 'tgps-user' folder). Run from inside TradersGPS repo.")


def _user_root(repo: Path) -> Path:
    return repo / "tgps-user"


def _cfg_path(user_root: Path) -> Path:
    return user_root / "config" / "lcl.user.yml"


def _ledger_path(user_root: Path) -> Path:
    return user_root / "ledger" / "lcl.ledger.duckdb"


def _sellput_outdir(user_root: Path) -> Path:
    return user_root / "output" / "sellput"


def _load_user_yml(p: Path) -> dict:
    if not p.exists():
        raise SystemExit(f"❌ Missing config: {p}")
    return yaml.safe_load(p.read_text()) or {}


def _sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _bot(cfg: dict, module: str) -> dict:
    b = cfg.get(f"{module}_bot")
    if isinstance(b, dict):
        return b
    b = cfg.get("sellput_bot")
    return b if isinstance(b, dict) else {}


def _num(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None


def _tick_size_from_cfg(bot: dict) -> float:
    pricing = bot.get("pricing") if isinstance(bot.get("pricing"), dict) else {}
    orders = bot.get("orders") if isinstance(bot.get("orders"), dict) else {}
    entry = bot.get("entry") if isinstance(bot.get("entry"), dict) else {}
    return float(
        _num((pricing or {}).get("tick_size"))
        or _num((orders or {}).get("tick_size"))
        or _num((entry or {}).get("rounding_tick"))
        or 0.05
    )


def _limit_default_pct_from_cfg(bot: dict) -> float:
    pricing = bot.get("pricing") if isinstance(bot.get("pricing"), dict) else {}
    orders = bot.get("orders") if isinstance(bot.get("orders"), dict) else {}
    entry = bot.get("entry") if isinstance(bot.get("entry"), dict) else {}
    ws = entry.get("wide_spread_rule") if isinstance(entry.get("wide_spread_rule"), dict) else {}
    return float(
        _num((pricing or {}).get("limit_default_pct_of_ask"))
        or _num((orders or {}).get("limit_default_pct_of_ask"))
        or _num((ws or {}).get("lmin_ask_pct"))
        or 0.90
    )


def _wide_thresholds_from_cfg(bot: dict) -> Tuple[float, float]:
    entry = bot.get("entry") if isinstance(bot.get("entry"), dict) else {}
    ws = entry.get("wide_spread_rule") if isinstance(entry.get("wide_spread_rule"), dict) else {}
    pct_thr = float(_num((ws or {}).get("pct_threshold")) or 0.25)
    abs_thr = float(_num((ws or {}).get("abs_threshold")) or 0.25)
    return pct_thr, abs_thr


# ----------------------------
# DuckDB attach
# ----------------------------
def _connect_attached(ledger: Path) -> duckdb.DuckDBPyConnection:
    con = duckdb.connect(":memory:")
    con.execute(f"ATTACH '{ledger.as_posix()}' AS ledger;")
    return con


def _colset(con: duckdb.DuckDBPyConnection, table: str) -> set[str]:
    rows = con.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_catalog='ledger' AND table_schema='lcl' AND table_name=?
        """,
        [table],
    ).fetchall()
    return {r[0] for r in rows}


def _ensure_actions_columns(con: duckdb.DuckDBPyConnection) -> None:
    wanted: Dict[str, str] = {
        "action_id": "VARCHAR",
        "created_at": "TIMESTAMP",
        "updated_at": "TIMESTAMP",
        "user_name": "VARCHAR",
        "module": "VARCHAR",
        "action": "VARCHAR",
        "status": "VARCHAR",
        "idea_id": "VARCHAR",
        "ticker": "VARCHAR",
        "expiry": "DATE",
        "strike": "DOUBLE",
        "qty": "DOUBLE",
        "price_override": "DOUBLE",
        "notes": "VARCHAR",
        "policy_decision": "VARCHAR",
        "policy_reasons": "VARCHAR",
        "action_ts": "TIMESTAMP",

        "strike_found": "DOUBLE",
        "qty_override": "DOUBLE",
        "limit_price_override": "DOUBLE",
        "user_notes": "VARCHAR",
        "source_queue_path": "VARCHAR",
        "source_queue_sha256": "VARCHAR",
        "source_row_index": "INTEGER",
        "action_hash": "VARCHAR",

        "entry_order_type": "VARCHAR",
        "duration": "VARCHAR",
        "attach_exit": "VARCHAR",
        "exit_mode": "VARCHAR",
        "tp_limit": "DOUBLE",
        "stop_type": "VARCHAR",
        "stop_price": "DOUBLE",
        "stop_limit_price": "DOUBLE",

        "q_bid": "DOUBLE",
        "q_ask": "DOUBLE",
        "q_mid": "DOUBLE",
        "q_spread_pct": "DOUBLE",
        "q_spread_wide": "BOOLEAN",
        "q_lmin": "DOUBLE",
        "q_defaulted_limit": "BOOLEAN",
    }

    cols = _colset(con, "actions")
    for c, typ in wanted.items():
        if c not in cols:
            con.execute(f"ALTER TABLE ledger.lcl.actions ADD COLUMN {c} {typ};")

    try:
        con.execute("CREATE UNIQUE INDEX IF NOT EXISTS ux_actions_action_hash ON ledger.lcl.actions(action_hash);")
    except Exception:
        pass

    con.execute(
        """
        CREATE VIEW IF NOT EXISTS ledger.lcl.v_actions_latest AS
        SELECT * EXCLUDE(rn)
        FROM (
            SELECT
                *,
                row_number() OVER (
                    PARTITION BY module, ticker, expiry, strike_found
                    ORDER BY action_ts DESC, created_at DESC
                ) AS rn
            FROM ledger.lcl.actions
        )
        WHERE rn = 1;
        """
    )


# ----------------------------
# Excel IO
# ----------------------------
def _latest_edited_queue(outdir: Path) -> Path:
    files = sorted(outdir.glob("*_sellput_queue_edit.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise SystemExit(
            f"❌ No edited queue files (*_sellput_queue_edit.xlsx) found in: {outdir}\n"
            f"Run: python -m scripts.tgps_user.sellput_day prep"
        )
    return files[0]


def _as_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def _as_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "")
    try:
        return float(s)
    except Exception:
        return None


def _as_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def _sheet_priority(wb: openpyxl.Workbook) -> List[str]:
    priority = ["ALL", "REVIEW", "WATCH", "SKIP"]
    names = list(wb.sheetnames)
    ordered = [n for n in priority if n in names]
    if ordered:
        return ordered
    try:
        return [wb.active.title]
    except Exception:
        return names[:1]


def _read_queue_rows(xlsx: Path) -> Tuple[List[str], List[dict]]:
    wb = _load_workbook_safe(xlsx, data_only=True)

    rows: List[dict] = []
    seen: set[tuple[str, date, float]] = set()
    headers_best: List[str] = []

    sheet_names = _sheet_priority(wb)

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        header_row = 1
        headers: List[str] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            if v is None:
                break
            headers.append(str(v).strip())

        if headers and not headers_best:
            headers_best = headers

        idx = {h: i for i, h in enumerate(headers)}

        required = ["Ticker", "Expiry", "Strike Found", "User Action"]
        if any(c not in idx for c in required):
            continue

        def get_cell(row: int, col_name: str) -> Any:
            if col_name not in idx:
                return None
            return ws.cell(row=row, column=idx[col_name] + 1).value

        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value in (None, ""):
                continue

            action = _as_str(get_cell(r, "User Action"))
            if not action:
                continue
            action_u = action.upper()
            if action_u not in VALID_ACTIONS:
                raise SystemExit(f"❌ Invalid 'User Action' at sheet={sheet_name} row={r}: {action_u}")

            ticker = _as_str(get_cell(r, "Ticker"))
            expiry = _as_date(get_cell(r, "Expiry"))
            strike_found = _as_float(get_cell(r, "Strike Found"))
            if not ticker or expiry is None or strike_found is None:
                raise SystemExit(f"❌ Missing Ticker/Expiry/Strike Found at sheet={sheet_name} row={r}")

            key = (ticker.strip().upper(), expiry, float(strike_found))
            if key in seen:
                continue
            seen.add(key)

            qty_v = _as_float(get_cell(r, "Qty"))
            if qty_v is None:
                qty_v = _as_float(get_cell(r, "Qty Override"))  # legacy fallback

            rec = {
                "excel_row": r,
                "ticker": ticker.strip().upper(),
                "expiry": expiry,
                "strike_found": float(strike_found),
                "action": action_u,

                "qty_override": qty_v,
                "limit_price_override": _as_float(get_cell(r, "Limit Price Override")),
                "user_notes": _as_str(get_cell(r, "User Notes")),

                "entry_order_type": (_as_str(get_cell(r, "Entry Order Type")) or "LIMIT").upper(),
                "duration": (_as_str(get_cell(r, "Duration")) or "DAY").upper(),
                "attach_exit": (_as_str(get_cell(r, "Attach Exit")) or "").upper(),
                "exit_mode": (_as_str(get_cell(r, "Exit Mode")) or "").upper(),
                "tp_limit": _as_float(get_cell(r, "TP Limit")),
                "stop_type": (_as_str(get_cell(r, "Stop Type")) or "").upper(),
                "stop_price": _as_float(get_cell(r, "Stop Price")),
                "stop_limit_price": _as_float(get_cell(r, "Stop Limit Price")),
            }
            rows.append(rec)

    try:
        wb.close()
    except Exception:
        pass

    if not rows and not headers_best:
        raise SystemExit("❌ No readable data sheets found in queue workbook.")

    return headers_best, rows

def _wait_for_stable_file(p: Path, *, stable_checks: int = 3, interval: float = 0.25, timeout: float = 6.0) -> None:
    """
    Wait until file's (size, mtime_ns) is stable for N consecutive checks.
    Helps when Excel/iCloud is mid-write or doing atomic replace.
    """
    t0 = time.time()
    last = None
    stable = 0

    while True:
        try:
            st = p.stat()
            cur = (st.st_size, getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))
        except FileNotFoundError:
            cur = None

        if cur is not None and last == cur and cur[0] > 0:
            stable += 1
            if stable >= stable_checks:
                return
        else:
            stable = 0
            last = cur

        if time.time() - t0 >= timeout:
            # best-effort: stop waiting and let load attempt happen
            return

        time.sleep(interval)


def _load_workbook_safe(xlsx: Path, *, data_only: bool = True, attempts: int = 10) -> openpyxl.Workbook:
    """
    Robustly open an .xlsx even if Excel/cloud sync is mid-save.
    Strategy: wait for stability, read into BytesIO, retry on BadZipFile.
    """
    xlsx = Path(xlsx)

    # Guard against Excel temp lock files being accidentally picked up
    if xlsx.name.startswith("~$"):
        raise SystemExit(f"❌ Refusing to open Excel lock file: {xlsx.name} (pick the real .xlsx)")

    last_err: Exception | None = None
    for i in range(attempts):
        try:
            _wait_for_stable_file(xlsx)

            # Read bytes first to avoid Excel atomic replace during openpyxl parsing
            with xlsx.open("rb") as f:
                blob = f.read()

            # A real xlsx is usually >> a few KB; tiny files are almost always mid-write
            if len(blob) < 4096:
                raise BadZipFile(f"Too small to be a complete .xlsx ({len(blob)} bytes)")

            bio = io.BytesIO(blob)
            return openpyxl.load_workbook(bio, data_only=data_only)

        except (BadZipFile, PermissionError, OSError) as e:
            last_err = e
            # backoff: 0.2s, 0.4s, ... capped
            time.sleep(min(1.2, 0.2 * (i + 1)))
            continue

    raise SystemExit(
        f"❌ Cannot read Excel workbook reliably (likely mid-save/cloud-sync).\n"
        f"File: {xlsx}\n"
        f"Last error: {last_err}"
    )

# ----------------------------
# Quote helpers + rounding
# ----------------------------
def _round_down_to_tick(x: float, tick: float) -> float:
    if tick <= 0:
        return float(x)
    return (int((float(x) + 1e-12) // tick)) * tick


def _round_up_to_tick(x: float, tick: float) -> float:
    if tick <= 0:
        return float(x)
    return (int(((float(x) - 1e-12) + tick - 1e-15) // tick) + 1) * tick if (float(x) % tick) > 1e-12 else float(x)


def _sanitize_underlying(sym: str) -> str:
    s = (sym or "").strip().upper()
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s


def _occ_option_symbol(underlying: str, expiry: date, right: str, strike: float) -> str:
    root = _sanitize_underlying(underlying)
    root = (root[:6]).ljust(6)
    yymmdd = expiry.strftime("%y%m%d")
    r = right.upper()
    if r not in ("P", "C"):
        raise ValueError(f"Invalid right: {right}")
    strike_i = int(round(float(strike) * 1000.0))
    return f"{root}{yymmdd}{r}{strike_i:08d}"


def _make_rest_client() -> RestClient:
    store = TokenStore()
    auth = AuthClient(store)
    session = RestSession(auth)
    return RestClient(session)


def _call_quotes(rest: RestClient, names: List[str], *args: Any, **kwargs: Any) -> Any:
    quotes = getattr(rest, "quotes", None) or getattr(rest, "market", None) or getattr(rest, "marketdata", None)
    if quotes is None:
        return None
    for n in names:
        fn = getattr(quotes, n, None)
        if callable(fn):
            return fn(*args, **kwargs)
    return None


def _fetch_option_bid_ask(option_symbol: str) -> Tuple[Optional[float], Optional[float]]:
    try:
        rest = _make_rest_client()
    except Exception:
        return None, None

    try:
        resp = _call_quotes(rest, ["get", "quote", "get_quote", "get_quotes", "quotes"], option_symbol)
        if resp is None:
            return None, None

        data = resp.get("data") if isinstance(resp, dict) and "data" in resp else resp

        if isinstance(data, dict):
            if "bid" in data or "ask" in data:
                return _as_float(data.get("bid")), _as_float(data.get("ask"))

            if option_symbol in data and isinstance(data[option_symbol], dict):
                d = data[option_symbol]
                return _as_float(d.get("bid")), _as_float(d.get("ask"))

        return None, None
    except Exception:
        return None, None


def _spread_pct(bid: float, ask: float) -> Optional[float]:
    mid = (bid + ask) / 2.0
    if mid <= 0:
        return None
    return (ask - bid) / mid


# ----------------------------
# Validation rules
# ----------------------------
@dataclass
class ValidatedRow:
    excel_row: int
    action: str
    ticker: str
    expiry: date
    strike_found: float

    qty: int
    entry_order_type: str
    duration: str
    limit_price: Optional[float]

    attach_exit: str
    exit_mode: str
    tp_limit: Optional[float]
    stop_type: str
    stop_price: Optional[float]
    stop_limit_price: Optional[float]

    user_notes: Optional[str]

    q_bid: Optional[float]
    q_ask: Optional[float]
    q_mid: Optional[float]
    q_spread_pct: Optional[float]
    q_spread_wide: Optional[bool]
    q_lmin: Optional[float]
    q_defaulted_limit: bool

    blockers: List[str]
    warnings: List[str]


def _normalize_duration(d: str) -> str:
    du = (d or "DAY").strip().upper()
    if du in ("GTC", "GOOD_TILL_CANCEL"):
        return "GOOD_TILL_CANCEL"
    return "DAY"


def _is_int_like(x: float) -> bool:
    return abs(x - round(x)) < 1e-9


def _validate_rows(
    rows: List[dict],
    *,
    module: str,
    pct_threshold: float,
    abs_threshold: float,
    tick_size: float,
    limit_default_pct_of_ask: float,
    default_qty: int,
) -> List[ValidatedRow]:
    out: List[ValidatedRow] = []

    for rec in rows:
        blockers: List[str] = []
        warnings: List[str] = []

        action = rec["action"]
        ticker = rec["ticker"]
        expiry = rec["expiry"]
        strike = float(rec["strike_found"])
        excel_row = int(rec["excel_row"])

        # Qty
        q_raw = rec.get("qty_override")
        if q_raw is None:
            qty = int(default_qty or 1)
        else:
            qf = float(q_raw)
            if not _is_int_like(qf):
                blockers.append("Qty must be an integer (e.g., 1,2,3).")
                qty = 0
            else:
                qty = int(round(qf))
        if qty <= 0:
            blockers.append("Qty must be >= 1.")

        # Entry type
        entry = (rec.get("entry_order_type") or "LIMIT").strip().upper()
        if entry not in VALID_ENTRY_TYPES:
            blockers.append(f"Invalid Entry Order Type: {entry}")
            entry = "LIMIT"

        duration = _normalize_duration(rec.get("duration") or "DAY")
        if duration not in VALID_DURATIONS:
            blockers.append(f"Invalid Duration: {duration}")

        attach_exit = (rec.get("attach_exit") or "").strip().upper()
        exit_mode = (rec.get("exit_mode") or "").strip().upper()
        stop_type = (rec.get("stop_type") or "").strip().upper()

        if action != "EXECUTE":
            attach_exit = "NO"
            exit_mode = "NONE"
            stop_type = "STOP_MARKET"
            tp_limit = None
            stop_price = None
            stop_limit = None
        else:
            if attach_exit in ("", None):
                attach_exit = "YES"
            if attach_exit != "YES":
                blockers.append("Attach Exit must be YES when User Action is EXECUTE.")
                attach_exit = "YES"

            if exit_mode in ("", None):
                exit_mode = "NONE"
            if exit_mode not in VALID_EXIT_MODES:
                blockers.append(f"Invalid Exit Mode: {exit_mode}")
                exit_mode = "NONE"

            if stop_type in ("", None):
                stop_type = "STOP_MARKET"
            if stop_type not in VALID_STOP_TYPES:
                blockers.append(f"Invalid Stop Type: {stop_type}")
                stop_type = "STOP_MARKET"

            tp_limit = rec.get("tp_limit")
            stop_price = rec.get("stop_price")
            stop_limit = rec.get("stop_limit_price")

            def has_any_exit_numbers() -> bool:
                return (tp_limit is not None) or (stop_price is not None) or (stop_limit is not None)

            if exit_mode == "NONE":
                if has_any_exit_numbers():
                    blockers.append("Exit Mode is NONE but TP/Stop prices were provided.")
                tp_limit = None
                stop_price = None
                stop_limit = None

            elif exit_mode == "TP_ONLY":
                if tp_limit is None:
                    blockers.append("TP_ONLY requires TP Limit.")
                if stop_price is not None or stop_limit is not None:
                    blockers.append("TP_ONLY does not allow Stop fields.")

            elif exit_mode == "STOP_ONLY":
                if stop_price is None:
                    blockers.append("STOP_ONLY requires Stop Price.")
                if tp_limit is not None:
                    blockers.append("STOP_ONLY does not allow TP Limit (must be empty).")
                if stop_type == "STOP_LIMIT":
                    if stop_limit is None:
                        blockers.append("STOP_LIMIT requires Stop Limit Price.")
                    elif stop_price is not None and float(stop_limit) < float(stop_price):
                        blockers.append("STOP_LIMIT safety: Stop Limit Price must be >= Stop Price (buy-to-close).")
                else:
                    stop_limit = None

            else:  # OCO
                if tp_limit is None:
                    blockers.append("OCO requires TP Limit.")
                if stop_price is None:
                    blockers.append("OCO requires Stop Price.")
                if stop_type == "STOP_LIMIT":
                    if stop_limit is None:
                        blockers.append("OCO STOP_LIMIT requires Stop Limit Price.")
                    elif stop_price is not None and float(stop_limit) < float(stop_price):
                        blockers.append("OCO STOP_LIMIT safety: Stop Limit Price must be >= Stop Price.")
                else:
                    stop_limit = None

        # LIMIT price rule
        limit_in = rec.get("limit_price_override")
        limit_price = (float(limit_in) if limit_in is not None else None)

        q_bid = q_ask = q_mid = q_sp = None
        q_wide: Optional[bool] = None
        q_lmin: Optional[float] = None
        defaulted_limit = False

        if entry == "MARKET":
            if limit_price is not None:
                blockers.append("MARKET entry must have blank Limit Price Override.")
            limit_price = None

        else:  # LIMIT
            option_symbol = ""
            try:
                option_symbol = _occ_option_symbol(ticker, expiry, "P", strike)
            except Exception:
                option_symbol = ""

            bid, ask = (None, None)
            if option_symbol:
                bid, ask = _fetch_option_bid_ask(option_symbol)

            if bid is not None and ask is not None and ask >= bid and ask >= 0 and bid >= 0:
                q_bid = float(bid)
                q_ask = float(ask)
                q_mid = (q_bid + q_ask) / 2.0

                abs_spread = q_ask - q_bid
                sp = _spread_pct(q_bid, q_ask)
                q_sp = sp

                # wide if pct >= pct_threshold OR abs >= abs_threshold
                q_wide = bool((sp is not None and sp >= float(pct_threshold)) or (abs_spread >= float(abs_threshold)))

                # Lmin = round_down_to_tick(max(Mid, Ask*pct))
                q_lmin = _round_down_to_tick(max(q_mid, q_ask * float(limit_default_pct_of_ask)), tick_size)

                if limit_price is None:
                    if q_wide:
                        limit_price = float(q_lmin)
                    else:
                        limit_price = float(_round_up_to_tick(q_mid, tick_size))
                    defaulted_limit = True
                else:
                    # If wide, user-provided limit must be >= Lmin
                    if q_wide and q_lmin is not None:
                        if float(limit_price) + 1e-9 < float(q_lmin):
                            blockers.append(
                                f"LIMIT too low for wide spread: need L >= {q_lmin:.2f} "
                                f"(round_down_to_tick(max(Mid, Ask*{limit_default_pct_of_ask:.2f})))."
                            )

            else:
                if limit_price is None:
                    blockers.append("LIMIT missing and Ask unavailable: cannot auto-default Limit Price Override.")
                else:
                    warnings.append("Ask unavailable: cannot validate wide-spread rule vs Ask/Mid.")

            if limit_price is None:
                blockers.append("Missing entry price: LIMIT requires Limit Price Override.")
            elif float(limit_price) <= 0:
                blockers.append("Limit Price Override must be > 0.")

        out.append(
            ValidatedRow(
                excel_row=excel_row,
                action=action,
                ticker=ticker,
                expiry=expiry,
                strike_found=strike,
                qty=qty if qty > 0 else 0,
                entry_order_type=entry,
                duration=duration,
                limit_price=limit_price,
                attach_exit=attach_exit,
                exit_mode=exit_mode,
                tp_limit=None if tp_limit is None else float(tp_limit),
                stop_type=stop_type,
                stop_price=None if stop_price is None else float(stop_price),
                stop_limit_price=None if stop_limit is None else float(stop_limit),
                user_notes=(None if rec.get("user_notes") is None else str(rec.get("user_notes"))),
                q_bid=q_bid,
                q_ask=q_ask,
                q_mid=q_mid,
                q_spread_pct=q_sp,
                q_spread_wide=q_wide,
                q_lmin=q_lmin,
                q_defaulted_limit=bool(defaulted_limit),
                blockers=blockers,
                warnings=warnings,
            )
        )

    return out


def _make_action_hash(queue_sha: str, v: ValidatedRow) -> str:
    key = (
        f"{queue_sha}|{v.excel_row}|{v.action}|"
        f"{v.qty}|{v.entry_order_type}|{v.duration}|{v.limit_price}|"
        f"{v.attach_exit}|{v.exit_mode}|{v.tp_limit}|{v.stop_type}|{v.stop_price}|{v.stop_limit_price}|"
        f"{v.user_notes}"
    )
    return hashlib.sha256(key.encode("utf-8")).hexdigest()


# ----------------------------
# Commands
# ----------------------------
def _resolve_queue_path(queue_path: Optional[str]) -> Path:
    repo = _repo_root()
    user_root = _user_root(repo)
    outdir = _sellput_outdir(user_root)
    if queue_path:
        q = Path(queue_path).expanduser().resolve()
    else:
        q = _latest_edited_queue(outdir)
    if not q.exists():
        raise SystemExit(f"❌ Queue not found: {q}")
    return q


def cmd_validate(queue_path: Optional[str]) -> int:
    repo = _repo_root()
    user_root = _user_root(repo)
    cfg = _load_user_yml(_cfg_path(user_root))

    module = (cfg.get("ideas_source", {}) or {}).get("module", "sellput")
    bot = _bot(cfg, module)

    default_qty = int((cfg.get("overrides", {}) or {}).get("default_qty", 1))
    tick_size = _tick_size_from_cfg(bot)
    limit_default_pct_of_ask = _limit_default_pct_from_cfg(bot)
    pct_threshold, abs_threshold = _wide_thresholds_from_cfg(bot)

    q = _resolve_queue_path(queue_path)
    queue_sha = _sha256_file(q)
    _, rows = _read_queue_rows(q)

    validated = _validate_rows(
        rows,
        module=module,
        pct_threshold=pct_threshold,
        abs_threshold=abs_threshold,
        tick_size=tick_size,
        limit_default_pct_of_ask=limit_default_pct_of_ask,
        default_qty=default_qty,
    )

    exec_rows = [v for v in validated if v.action == "EXECUTE"]
    blocked_exec = [v for v in exec_rows if v.blockers]
    ok_exec = [v for v in exec_rows if not v.blockers]

    print(f"✅ validate: queue={q.name}")
    print(f"  source_queue_sha256: {queue_sha[:12]}…")
    print(f"  default_qty:         {default_qty}")
    print(f"  tick_size:           {tick_size}")
    print(f"  wide_pct_threshold:  {pct_threshold}")
    print(f"  wide_abs_threshold:  {abs_threshold}")
    print(f"  ask_default_pct:     {limit_default_pct_of_ask}")
    print(f"  rows_with_actions:   {len(validated)}")
    print(f"  EXECUTE:             {len(exec_rows)}  OK={len(ok_exec)}  BLOCKED={len(blocked_exec)}")

    if blocked_exec:
        print("\n[BLOCKED EXECUTE ROWS]")
        for v in blocked_exec[:40]:
            print(f" - row={v.excel_row} {v.ticker} {v.expiry} P{v.strike_found:g} :: " + "; ".join(v.blockers))
        if len(blocked_exec) > 40:
            print(f" ... ({len(blocked_exec) - 40} more)")

    defaulted = [v for v in exec_rows if v.q_defaulted_limit]
    if defaulted:
        print("\n[DEFAULTS APPLIED]")
        for v in defaulted[:40]:
            print(
                f" - row={v.excel_row} {v.ticker} {v.expiry} P{v.strike_found:g} "
                f"DEFAULT Limit={v.limit_price:.2f} "
                f"(bid={'' if v.q_bid is None else f'{v.q_bid:.2f}'} "
                f"ask={'' if v.q_ask is None else f'{v.q_ask:.2f}'} wide={v.q_spread_wide})"
            )

    warned = [v for v in exec_rows if v.warnings]
    if warned:
        print("\n[WARNINGS]")
        for v in warned[:40]:
            print(f" - row={v.excel_row} {v.ticker} :: " + "; ".join(v.warnings))

    return 2 if blocked_exec else 0


def cmd_import(queue_path: Optional[str]) -> int:
    repo = _repo_root()
    user_root = _user_root(repo)

    cfg = _load_user_yml(_cfg_path(user_root))
    module = (cfg.get("ideas_source", {}) or {}).get("module", "sellput")
    bot = _bot(cfg, module)

    default_qty = int((cfg.get("overrides", {}) or {}).get("default_qty", 1))
    tick_size = _tick_size_from_cfg(bot)
    limit_default_pct_of_ask = _limit_default_pct_from_cfg(bot)
    pct_threshold, abs_threshold = _wide_thresholds_from_cfg(bot)

    ledger = _ledger_path(user_root)
    if not ledger.exists():
        raise SystemExit(f"❌ Ledger not found: {ledger}")

    q = _resolve_queue_path(queue_path)
    queue_sha = _sha256_file(q)
    _, rows = _read_queue_rows(q)

    validated = _validate_rows(
        rows,
        module=module,
        pct_threshold=pct_threshold,
        abs_threshold=abs_threshold,
        tick_size=tick_size,
        limit_default_pct_of_ask=limit_default_pct_of_ask,
        default_qty=default_qty,
    )

    blocked_exec = [v for v in validated if v.action == "EXECUTE" and v.blockers]
    if blocked_exec:
        print("❌ import refused: BLOCKED EXECUTE rows exist. Fix Excel then re-run validate/import.")
        cmd_validate(str(q))
        return 2

    con = _connect_attached(ledger)
    _ensure_actions_columns(con)

    inserted = 0
    skipped_dupe = 0
    counts: Dict[str, int] = {}

    now_ts = datetime.now()

    for v in validated:
        action_hash = _make_action_hash(queue_sha, v)
        action_id = uuid.uuid4().hex

        try:
            con.execute(
                """
                INSERT INTO ledger.lcl.actions (
                    action_id, action_ts, module, action,
                    ticker, expiry, strike_found,
                    qty_override, limit_price_override, user_notes,
                    entry_order_type, duration, attach_exit,
                    exit_mode, tp_limit, stop_type, stop_price, stop_limit_price,
                    source_queue_path, source_queue_sha256, source_row_index,
                    action_hash, created_at,

                    q_bid, q_ask, q_mid, q_spread_pct, q_spread_wide, q_lmin, q_defaulted_limit
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    action_id,
                    now_ts,
                    module,
                    v.action,
                    v.ticker,
                    v.expiry,
                    v.strike_found,
                    float(v.qty),
                    (None if v.limit_price is None else float(v.limit_price)),
                    v.user_notes,
                    v.entry_order_type,
                    v.duration,
                    v.attach_exit,
                    v.exit_mode,
                    v.tp_limit,
                    v.stop_type,
                    v.stop_price,
                    v.stop_limit_price,
                    str(q),
                    queue_sha,
                    int(v.excel_row),
                    action_hash,
                    now_ts,
                    v.q_bid,
                    v.q_ask,
                    v.q_mid,
                    v.q_spread_pct,
                    v.q_spread_wide,
                    v.q_lmin,
                    bool(v.q_defaulted_limit),
                ],
            )

            inserted += 1
            counts[v.action] = counts.get(v.action, 0) + 1
        except Exception as e:
            msg = str(e).lower()
            if "duplicate" in msg or "constraint" in msg or "unique" in msg:
                skipped_dupe += 1
            else:
                raise

    con.close()

    print(f"✅ actions_capture.import: queue={q.name}")
    print(f"  source_queue_sha256: {queue_sha[:12]}…")
    print(f"  rows_with_actions:   {len(validated)}")
    print(f"  inserted:           {inserted}")
    print(f"  skipped_dupe:       {skipped_dupe}")
    print(f"  counts:             {counts}")
    return 0


def cmd_purge(queue_path: Optional[str], yes: bool) -> int:
    repo = _repo_root()
    user_root = _user_root(repo)
    ledger = _ledger_path(user_root)
    if not ledger.exists():
        raise SystemExit(f"❌ Ledger not found: {ledger}")

    q = _resolve_queue_path(queue_path)
    queue_sha = _sha256_file(q)

    con = _connect_attached(ledger)
    _ensure_actions_columns(con)

    n = con.execute(
        "SELECT count(*) FROM ledger.lcl.actions WHERE source_queue_sha256 = ?",
        [queue_sha],
    ).fetchone()[0]

    print(f"[purge] queue={q.name}")
    print(f"        source_queue_sha256={queue_sha}")
    print(f"        matching_actions={n}")

    if n == 0:
        con.close()
        return 0

    if not yes:
        print("DRY-RUN: add --yes to delete these rows.")
        con.close()
        return 0

    con.execute("DELETE FROM ledger.lcl.actions WHERE source_queue_sha256 = ?", [queue_sha])
    con.close()
    print("✅ purge done.")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_val = sub.add_parser("validate", help="Validate queue_edit (defaults + hard blocks)")
    p_val.add_argument("--queue", default=None)

    p_imp = sub.add_parser("import", help="Import actions from edited queue (validate first; hard-stop on blocked)")
    p_imp.add_argument("--queue", default=None)

    p_purge = sub.add_parser("purge", help="Delete actions rows for a queue (by source_queue_sha256)")
    p_purge.add_argument("--queue", default=None)
    p_purge.add_argument("--yes", action="store_true")

    args = ap.parse_args()

    if args.cmd == "validate":
        return cmd_validate(args.queue)
    if args.cmd == "import":
        return cmd_import(args.queue)
    if args.cmd == "purge":
        return cmd_purge(args.queue, bool(args.yes))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
