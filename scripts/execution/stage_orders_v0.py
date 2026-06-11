#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/execution/stage_orders_v0.py
Version: 0.1.1
Updated: 2025-12-16 (SGT)

PHASE 0 ONLY (DRY-RUN)
- Reads Excel Orders_V0 sheet
- Applies KISS contract validation (V0)
- Applies kill-switch + circuit breaker
- LOCK B: option symbol must come from option-chain cache lookup (fail-closed until wired)
- Generates Schwab Order JSON payload (not sent)
- Writes back ExecStatus/BlockReason/StagedAt/OrderJson
- Writes JSONL audit log

Sheet layout (LOCKED)
- B1 = GlobalKillSwitch (ON/OFF)
- Row 2 = headers
- Row 3+ = data
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

import openpyxl

# --- Optional YAML -----------------------------------------------------------
try:
    import yaml  # type: ignore
except Exception:  # pragma: no cover
    yaml = None


# ----------------------------
# Constants (LOCKED layout)
# ----------------------------
HEADER_ROW = 2
DATA_START_ROW = 3


# ----------------------------
# Helpers
# ----------------------------


def now_sgt_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm_str(v: Any) -> str:
    return ("" if v is None else str(v)).strip()


def as_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def as_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float) and v.is_integer():
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def is_yes(v: Any) -> bool:
    return norm_str(v).upper() == "YES"


def excel_date_to_ymd(v: Any) -> Optional[str]:
    """
    Accept:
    - datetime/date objects
    - YYYY-MM-DD string
    - dd/mm/yy, dd/mm/yyyy, d/m/yy
    - dd-mm-yy, dd-mm-yyyy, d-m-yy
    """
    if v is None:
        return None

    # openpyxl often gives datetime for real Excel dates
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return None

    s = norm_str(v)
    if not s:
        return None

    # Already ISO
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return s

    # Try common day-first formats
    s2 = s.replace(".", "/").replace("-", "/")
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(s2, fmt)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass

    # Try non-zero padded day/month (e.g. 9/1/26)
    parts = s2.split("/")
    if len(parts) == 3:
        try:
            d = int(parts[0])
            m = int(parts[1])
            y = int(parts[2])
            if y < 100:
                y += 2000
            dt = datetime(y, m, d)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return None

    return None


# ----------------------------
# Config
# ----------------------------


@dataclass
class ExecConfig:
    kill_switch: bool
    account_aliases: Dict[str, str]
    defaults_session: str
    defaults_duration: str
    defaults_order_type: str
    max_orders_per_run: int
    max_notional_per_order: float
    max_contracts_per_order: int
    sheet_name: str
    global_kill_cell: str
    log_dir: Path
    option_cache_dir: Path


def load_config(cfg_path: Path) -> ExecConfig:
    if not cfg_path.exists():
        raise FileNotFoundError(f"Config not found: {cfg_path}")

    raw: Dict[str, Any]
    if cfg_path.suffix.lower() in (".yaml", ".yml"):
        if yaml is None:
            raise RuntimeError(
                "PyYAML not installed. Use JSON config or: pip install pyyaml"
            )
        raw = yaml.safe_load(cfg_path.read_text()) or {}
    else:
        raw = json.loads(cfg_path.read_text())

    kill_switch = bool(raw.get("kill_switch", True))

    account_aliases = raw.get("account_aliases", {}) or {}
    defaults = raw.get("defaults", {}) or {}
    limits = raw.get("limits", {}) or {}
    excel = raw.get("excel", {}) or {}

    defaults_session = norm_str(defaults.get("session") or "NORMAL").upper()
    defaults_duration = norm_str(defaults.get("duration") or "DAY").upper()
    defaults_order_type = norm_str(defaults.get("orderType") or "LIMIT").upper()

    max_orders_per_run = int(limits.get("max_orders_per_run", 1))
    max_notional_per_order = float(limits.get("max_notional_per_order", 5000))
    max_contracts_per_order = int(limits.get("max_contracts_per_order", 1))

    sheet_name = norm_str(excel.get("sheet_name") or "Orders_V0")
    global_kill_cell = norm_str(excel.get("global_kill_cell") or "B1")

    log_dir = Path(raw.get("log_dir", "output/execution_logs")).expanduser()
    option_cache_dir = Path(
        raw.get("option_cache_dir", "data/options_cache")
    ).expanduser()

    return ExecConfig(
        kill_switch=kill_switch,
        account_aliases=account_aliases,
        defaults_session=defaults_session,
        defaults_duration=defaults_duration,
        defaults_order_type=defaults_order_type,
        max_orders_per_run=max_orders_per_run,
        max_notional_per_order=max_notional_per_order,
        max_contracts_per_order=max_contracts_per_order,
        sheet_name=sheet_name,
        global_kill_cell=global_kill_cell,
        log_dir=log_dir,
        option_cache_dir=option_cache_dir,
    )


# ----------------------------
# LOCK B: option symbol lookup hook
# ----------------------------


class OptionSymbolLookupError(Exception):
    pass


def _load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _strike_equal(a: float, b: float, tol: float = 1e-6) -> bool:
    return abs(float(a) - float(b)) <= tol


def resolve_option_symbol_from_cache(
    *,
    ticker: str,
    expiry_ymd: str,
    strike: float,
    right: str,
    cfg: ExecConfig,
) -> str:
    """
    LOCK B (V0):
    Resolve exact Schwab option symbol from:
    data/options_cache/<TICKER>/chains/<YYYY-MM-DD>/puts.json|calls.json
    """
    ticker = ticker.upper()
    right = right.upper()
    if right not in ("PUT", "CALL"):
        raise OptionSymbolLookupError(f"Unsupported right: {right}")

    base = cfg.option_cache_dir / ticker / "chains" / expiry_ymd
    if not base.exists():
        raise OptionSymbolLookupError(
            f"No option chain folder for {ticker} {expiry_ymd} at {base}"
        )

    fname = "puts.json" if right == "PUT" else "calls.json"
    path = base / fname
    if not path.exists():
        raise OptionSymbolLookupError(f"Missing {fname} for {ticker} {expiry_ymd}")

    data = _load_json(path)

    if not isinstance(data, list):
        raise OptionSymbolLookupError(f"Unexpected schema in {path} (expected list)")

    for c in data:
        if not isinstance(c, dict):
            continue

        st = c.get("strike") or c.get("strikePrice")
        try:
            stf = float(st)
        except Exception:
            continue

        if not _strike_equal(stf, strike):
            continue

        sym = c.get("contract") or c.get("optionSymbol") or c.get("symbol")

        if not sym:
            continue

        return str(sym)

    raise OptionSymbolLookupError(
        f"Contract not found in {path.name} for strike {strike}"
    )


# ----------------------------
# Excel staging
# ----------------------------

REQUIRED_COLUMNS = [
    "Approve",
    "Action",
    "ClientOrderTag",
    "AccountAlias",
    "Ticker",
    "OptionType",
    "Side",
    "Expiry",
    "Strike",
    "Contracts",
    "OrderType",
    "LimitPrice",
    "TimeInForce",
    "Session",
    "ExecStatus",
    "BlockReason",
    "StagedAt",
    "OrderJson",
]

ALLOWED = {
    "OptionType": {"PUT"},
    "Side": {"SELL_TO_OPEN"},
    "Action": {"SUBMIT"},
    "OrderType": {"LIMIT"},
    "TimeInForce": {"DAY"},
    "Session": {"NORMAL"},
}

FIXED_JSON = {
    "orderStrategyType": "SINGLE",
    "orderLegType": "OPTION",
    "instruction": "SELL_TO_OPEN",
    "positionEffect": "OPENING",
    "assetType": "OPTION",
}


def ensure_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=HEADER_ROW, column=col).value
        name = norm_str(val)
        if name:
            headers[name] = col

    next_col = ws.max_column + 1
    for name in REQUIRED_COLUMNS:
        if name not in headers:
            ws.cell(row=HEADER_ROW, column=next_col).value = name
            headers[name] = next_col
            next_col += 1

    return headers


def get_cell(ws, row: int, headers: Dict[str, int], name: str) -> Any:
    return ws.cell(row=row, column=headers[name]).value


def set_cell(ws, row: int, headers: Dict[str, int], name: str, value: Any) -> None:
    ws.cell(row=row, column=headers[name]).value = value


def build_order_json(
    *,
    option_symbol: str,
    limit_price: float,
    contracts: int,
    tag: str,
    session: str,
    duration: str,
    order_type: str,
) -> Dict[str, Any]:
    return {
        "session": session,
        "duration": duration,
        "orderType": order_type,
        "price": float(limit_price),
        "orderStrategyType": FIXED_JSON["orderStrategyType"],
        "tag": tag,
        "orderLegCollection": [
            {
                "orderLegType": FIXED_JSON["orderLegType"],
                "instruction": FIXED_JSON["instruction"],
                "positionEffect": FIXED_JSON["positionEffect"],
                "quantity": int(contracts),
                "instrument": {
                    "symbol": option_symbol,
                    "assetType": FIXED_JSON["assetType"],
                },
            }
        ],
    }


def stage_orders(*, xlsx_path: Path, cfg: ExecConfig) -> Tuple[int, int, int, Path]:
    if norm_str(os.environ.get("TGPS_EXEC_DISABLE")).strip() == "1":
        raise RuntimeError("KILL_SWITCH_ENV_ON (TGPS_EXEC_DISABLE=1)")
    if cfg.kill_switch:
        raise RuntimeError("KILL_SWITCH_CONFIG_ON (kill_switch=true)")

    wb = openpyxl.load_workbook(xlsx_path)
    if cfg.sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {cfg.sheet_name}. Found: {wb.sheetnames}")
    ws = wb[cfg.sheet_name]
    headers = ensure_headers(ws)

    gks = norm_str(ws[cfg.global_kill_cell].value).upper()
    if gks == "OFF":
        raise RuntimeError(f"KILL_SWITCH_EXCEL_OFF ({cfg.global_kill_cell}=OFF)")

    cfg.log_dir.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = cfg.log_dir / f"stage_orders_v0_{run_id}.jsonl"

    n_seen = 0
    n_staged = 0
    n_blocked = 0
    staged_this_run = 0

    for row in range(DATA_START_ROW, ws.max_row + 1):
        approve = get_cell(ws, row, headers, "Approve")
        if not is_yes(approve):
            continue

        n_seen += 1

        def block(reason: str):
            nonlocal n_blocked
            set_cell(ws, row, headers, "ExecStatus", "BLOCKED")
            set_cell(ws, row, headers, "BlockReason", reason)
            set_cell(ws, row, headers, "StagedAt", now_sgt_iso())
            set_cell(ws, row, headers, "OrderJson", "")
            n_blocked += 1
            with log_path.open("a", encoding="utf-8") as f:
                f.write(
                    json.dumps(
                        {
                            "ts": now_sgt_iso(),
                            "row": row,
                            "clientTag": norm_str(
                                get_cell(ws, row, headers, "ClientOrderTag")
                            ),
                            "status": "BLOCKED",
                            "reason": reason,
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                )

        action = norm_str(get_cell(ws, row, headers, "Action")).upper()
        if action not in ALLOWED["Action"]:
            block(f"Action must be SUBMIT (V0). Got: {action or '<blank>'}")
            continue

        acct_alias = norm_str(get_cell(ws, row, headers, "AccountAlias"))
        client_tag = norm_str(get_cell(ws, row, headers, "ClientOrderTag"))
        ticker = norm_str(get_cell(ws, row, headers, "Ticker")).upper()
        opt_type = norm_str(get_cell(ws, row, headers, "OptionType")).upper()
        side = norm_str(get_cell(ws, row, headers, "Side")).upper()
        expiry = excel_date_to_ymd(get_cell(ws, row, headers, "Expiry"))
        strike = as_float(get_cell(ws, row, headers, "Strike"))
        contracts = as_int(get_cell(ws, row, headers, "Contracts"))

        order_type = norm_str(get_cell(ws, row, headers, "OrderType")).upper()
        limit_price = as_float(get_cell(ws, row, headers, "LimitPrice"))
        tif = norm_str(get_cell(ws, row, headers, "TimeInForce")).upper()
        session_cell = norm_str(get_cell(ws, row, headers, "Session")).upper()

        if not client_tag:
            block("ClientOrderTag is required and must be unique.")
            continue
        if not acct_alias:
            block("AccountAlias is required (expected: Main).")
            continue
        if acct_alias not in cfg.account_aliases:
            block(f"AccountAlias '{acct_alias}' not found in config account_aliases.")
            continue

        if not ticker:
            block("Ticker is required.")
            continue
        if opt_type not in ALLOWED["OptionType"]:
            block(f"OptionType must be PUT (V0). Got: {opt_type or '<blank>'}")
            continue
        if side not in ALLOWED["Side"]:
            block(f"Side must be SELL_TO_OPEN (V0). Got: {side or '<blank>'}")
            continue
        if not expiry:
            block("Expiry invalid. Use YYYY-MM-DD or day-first like 23/1/26.")
            continue
        if strike is None or strike <= 0:
            block("Strike must be a positive number.")
            continue
        if contracts is None or contracts < 1:
            block("Contracts must be an integer >= 1.")
            continue

        if order_type not in ALLOWED["OrderType"]:
            block(f"OrderType must be LIMIT (V0). Got: {order_type or '<blank>'}")
            continue
        if tif not in ALLOWED["TimeInForce"]:
            block(f"TimeInForce must be DAY (V0). Got: {tif or '<blank>'}")
            continue
        if limit_price is None or limit_price <= 0:
            block("LimitPrice must be a positive number.")
            continue

        session = session_cell or cfg.defaults_session
        if session not in ALLOWED["Session"]:
            block(f"Session must be NORMAL (V0). Got: {session}")
            continue

        if contracts > cfg.max_contracts_per_order:
            block(
                f"Contracts exceed max_contracts_per_order={cfg.max_contracts_per_order}."
            )
            continue

        notional = float(strike) * 100.0 * float(contracts)
        if notional > cfg.max_notional_per_order:
            block(
                f"Notional ${notional:,.2f} exceeds max_notional_per_order=${cfg.max_notional_per_order:,.2f}."
            )
            continue

        if staged_this_run + 1 > cfg.max_orders_per_run:
            block(f"max_orders_per_run={cfg.max_orders_per_run} reached.")
            continue

        try:
            option_symbol = resolve_option_symbol_from_cache(
                ticker=ticker,
                expiry_ymd=expiry,
                strike=float(strike),
                right="PUT",
                cfg=cfg,
            )
        except OptionSymbolLookupError as e:
            block(str(e))
            continue
        except Exception as e:
            block(f"Option symbol lookup error: {e}")
            continue

        order_json = build_order_json(
            option_symbol=option_symbol,
            limit_price=float(limit_price),
            contracts=int(contracts),
            tag=client_tag,
            session=session,
            duration=cfg.defaults_duration,
            order_type=cfg.defaults_order_type,
        )

        set_cell(ws, row, headers, "ExecStatus", "STAGED")
        set_cell(ws, row, headers, "BlockReason", "")
        set_cell(ws, row, headers, "StagedAt", now_sgt_iso())
        set_cell(
            ws, row, headers, "OrderJson", json.dumps(order_json, ensure_ascii=False)
        )

        staged_this_run += 1
        n_staged += 1

        with log_path.open("a", encoding="utf-8") as f:
            f.write(
                json.dumps(
                    {
                        "ts": now_sgt_iso(),
                        "row": row,
                        "clientTag": client_tag,
                        "status": "STAGED",
                        "ticker": ticker,
                        "expiry": expiry,
                        "strike": strike,
                        "contracts": contracts,
                        "limitPrice": limit_price,
                        "notional": notional,
                        "optionSymbol": option_symbol,
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )

    wb.save(xlsx_path)
    return n_seen, n_staged, n_blocked, log_path


def main():
    ap = argparse.ArgumentParser(
        description="Phase 0: Stage Schwab orders (dry-run only)."
    )
    ap.add_argument(
        "--xlsx",
        required=True,
        help="Path to <date>_user_sellput_ideas.xlsx (modified in-place).",
    )
    ap.add_argument(
        "--config",
        default="config/schwab_exec.json",
        help="Path to exec config JSON/YAML.",
    )
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    cfg_path = Path(args.config).expanduser()

    cfg = load_config(cfg_path)

    try:
        n_seen, n_staged, n_blocked, log_path = stage_orders(
            xlsx_path=xlsx_path, cfg=cfg
        )
    except Exception as e:
        print(f"✖ Phase 0 staging stopped: {e}")
        sys.exit(2)

    print("✓ Phase 0 staging complete")
    print(f"  Approved rows seen: {n_seen}")
    print(f"  STAGED:             {n_staged}")
    print(f"  BLOCKED:            {n_blocked}")
    print(f"  Log:                {log_path}")
    sys.exit(0)


if __name__ == "__main__":
    main()
