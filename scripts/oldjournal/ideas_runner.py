#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/journal/ideas_runner.py

Version: 0.1.0
Updated: 2026-01-08 (SGT)

Purpose
-------
Read your "sell put ideas" Excel, apply a simple gating/highlight ruleset,
and generate ready-to-run `oms_cli place-trigger-oco ...` command lines.

Design philosophy
-----------------
- Keep `oms_cli.py` as the low-level tool (place/replace/cancel).
- Use this runner as the orchestration layer:
  - Defaults live in ONE place (top of file + optional config JSON)
  - Produces an output Excel for review + a .sh file of commands
  - Does NOT submit orders unless you explicitly ask it to execute

Notes
-----
- Works with columns in your ideas workbook (like your uploaded file):
  Ticker, Win Rate (%), Strike Found, Ideal Strike, Underlying, Expiry, DTE,
  Last, Open Interest, Annualised Simple (%)
- If an "Earnings" column exists, it will be used to flag earnings risk.

Usage (typical)
---------------
# 1) Review output + generate commands (dry-run, default)
python -m scripts.journal.ideas_runner --file "/path/to/ideas.xlsx" --outdir "/path/to/output"

# 2) Include REVIEW ideas in generated .sh
python -m scripts.journal.ideas_runner --file "/path/to/ideas.xlsx" --include-review

# 3) Execute commands (adds --submit and runs them)
python -m scripts.journal.ideas_runner --file "/path/to/ideas.xlsx" --execute --include-review
"""

from __future__ import annotations

import argparse
import json
import math
import os
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


SGT = timezone(timedelta(hours=8))


# ----------------------------
# Defaults (easy to tweak)
# ----------------------------
DEFAULTS: Dict[str, Any] = {
    # gating thresholds
    "min_open_interest_review": 50,      # below this => REVIEW (liquidity warning)
    "min_open_interest_skip": 10,        # below this => SKIP (too illiquid)
    "min_annualised_roi_skip": 10.0,     # below this => SKIP (too low)
    "min_win_rate_review": 60.0,         # below this => REVIEW
    "min_win_rate_skip": 45.0,           # below this => SKIP
    "min_dte_review": 7,                 # below this => REVIEW (gamma/assignment risk)
    "min_dte_skip": 3,                   # below this => SKIP

    # earnings risk flags (only if Earnings column exists)
    "earnings_risk_weeks": 8,            # flag if earnings within this many weeks AND before expiry
    "earnings_missing_policy": "ignore", # ignore|review|skip

    # order construction
    "right": "PUT",                      # this runner targets sell puts
    "entry_type": "LIMIT",               # safest for credit entry
    "profit_take_frac": 0.50,            # buy back at 50% of entry credit
    "stop_loss_mult": 2.00,              # stop triggers at 2x entry credit
    "use_stop_limit": False,             # if True, also set stop-limit price
    "stop_limit_mult": 2.20,             # stop-limit price = entry_credit * stop_limit_mult
    "qty": 1,
    "session": "NORMAL",
    "duration": "GOOD_TILL_CANCEL",
    "complex_type": "NONE",

    # execution
    "oms_cli_module": "scripts.journal.oms_cli",  # used with `python -m ...`
}


# ----------------------------
# Helpers
# ----------------------------
def _today_sgt() -> datetime:
    return datetime.now(SGT)


def _as_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, float) and math.isnan(x):
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return None


def _as_int(x: Any) -> Optional[int]:
    f = _as_float(x)
    if f is None:
        return None
    return int(round(f))


def _parse_date(x: Any) -> Optional[datetime]:
    """Best-effort parse to a datetime (date at midnight in SGT)."""
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.astimezone(SGT) if x.tzinfo else x.replace(tzinfo=SGT)
    try:
        ts = pd.to_datetime(x, errors="coerce")
        if pd.isna(ts):
            return None
        dtv = ts.to_pydatetime()
        return dtv.astimezone(SGT) if dtv.tzinfo else dtv.replace(tzinfo=SGT)
    except Exception:
        return None


def _tick_size(price: float) -> Decimal:
    # Common US options tick rules: < 3.00 => $0.01 else $0.05
    return Decimal("0.01") if Decimal(str(price)) < Decimal("3") else Decimal("0.05")


def _round_to_tick(price: float) -> float:
    p = Decimal(str(price))
    tick = _tick_size(price)
    n = (p / tick).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return float((n * tick).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def build_osi_symbol(root: str, expiry: datetime, right: str, strike: float) -> str:
    """
    OSI format used by Schwab:
      ROOT padded to 6 chars (spaces)
      YYMMDD
      C/P
      STRIKE * 1000 as 8 digits

    Example: KTOS  260116P00092500
    """
    r = (root or "").upper().strip()
    r6 = (r + "      ")[:6]
    yymmdd = expiry.strftime("%y%m%d")
    cp = "P" if right.upper().startswith("P") else "C"
    strike_int = int(round(float(strike) * 1000))
    strike_str = f"{strike_int:08d}"
    return f"{r6}{yymmdd}{cp}{strike_str}"


def _join_reasons(reasons: List[str]) -> str:
    return "; ".join(dict.fromkeys([r for r in reasons if r]))  # stable unique


@dataclass
class IdeaRow:
    ticker: str
    expiry: datetime
    strike: float
    underlying: float
    last: float
    open_interest: int
    dte: int
    win_rate: Optional[float]
    annualised_roi: Optional[float]
    earnings: Optional[datetime]


def load_ideas_excel(path: str, sheet: Optional[str] = None) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    sh = sheet or xls.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sh)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _extract_idea(row: pd.Series) -> Optional[IdeaRow]:
    ticker = str(row.get("Ticker") or "").strip().upper()
    if not ticker:
        return None

    expiry_dt = _parse_date(row.get("Expiry"))
    if not expiry_dt:
        return None

    strike = _as_float(row.get("Strike Found")) or _as_float(row.get("Ideal Strike"))
    underlying = _as_float(row.get("Underlying"))
    last = _as_float(row.get("Last"))
    oi = _as_int(row.get("Open Interest")) or 0
    dte = _as_int(row.get("DTE")) or 0

    if strike is None or underlying is None or last is None:
        return None

    win_rate = _as_float(row.get("Win Rate (%)"))
    annual = _as_float(row.get("Annualised Simple (%)"))

    # Excel % cells often arrive as fractions (e.g. 0.82 for 82%).
    if win_rate is not None and win_rate <= 1.0:
        win_rate *= 100.0
    if annual is not None and annual <= 1.0:
        annual *= 100.0

    earnings_dt = _parse_date(row.get("Earnings")) if "Earnings" in row.index else None

    return IdeaRow(
        ticker=ticker,
        expiry=expiry_dt,
        strike=float(strike),
        underlying=float(underlying),
        last=float(last),
        open_interest=int(oi),
        dte=int(dte),
        win_rate=None if win_rate is None else float(win_rate),
        annualised_roi=None if annual is None else float(annual),
        earnings=earnings_dt,
    )


def decide_and_build_command(
    idea: IdeaRow,
    cfg: Dict[str, Any],
    *,
    account: Optional[str],
    force: bool,
) -> Tuple[str, str, str]:
    """
    Return (decision, reasons, command)

    decision:
      OK      -> include by default
      REVIEW  -> warnings (include only with --include-review or --force)
      SKIP    -> hard stop (include only with --include-skip AND --force)
    """
    reasons: List[str] = []
    decision = "OK"

    # --- Liquidity ---
    if idea.open_interest < int(cfg["min_open_interest_skip"]):
        decision = "SKIP"
        reasons.append(f"OI<{cfg['min_open_interest_skip']}")
    elif idea.open_interest < int(cfg["min_open_interest_review"]):
        decision = "REVIEW"
        reasons.append(f"OI<{cfg['min_open_interest_review']}")

    # --- ROI ---
    if idea.annualised_roi is not None and idea.annualised_roi < float(cfg["min_annualised_roi_skip"]):
        decision = "SKIP"
        reasons.append(f"ROI<{cfg['min_annualised_roi_skip']}%")

    # --- Win rate ---
    if idea.win_rate is not None:
        if idea.win_rate < float(cfg["min_win_rate_skip"]):
            decision = "SKIP"
            reasons.append(f"WinRate<{cfg['min_win_rate_skip']}%")
        elif idea.win_rate < float(cfg["min_win_rate_review"]) and decision != "SKIP":
            decision = "REVIEW"
            reasons.append(f"WinRate<{cfg['min_win_rate_review']}%")

    # --- DTE ---
    if idea.dte < int(cfg["min_dte_skip"]):
        decision = "SKIP"
        reasons.append(f"DTE<{cfg['min_dte_skip']}")
    elif idea.dte < int(cfg["min_dte_review"]) and decision != "SKIP":
        decision = "REVIEW"
        reasons.append(f"DTE<{cfg['min_dte_review']}")

    # --- Earnings risk ---
    if idea.earnings:
        today = _today_sgt().date()
        e_date = idea.earnings.date()
        exp_date = idea.expiry.date()
        if e_date >= today and e_date <= exp_date:
            weeks = float(cfg["earnings_risk_weeks"])
            if (e_date - today).days <= int(weeks * 7):
                if decision != "SKIP":
                    decision = "REVIEW"
                reasons.append(f"Earnings({e_date.isoformat()})<=Expiry")
    else:
        pol = str(cfg.get("earnings_missing_policy", "ignore")).lower()
        if pol == "review" and decision != "SKIP":
            decision = "REVIEW"
            reasons.append("EarningsMissing")
        elif pol == "skip":
            decision = "SKIP"
            reasons.append("EarningsMissing")

    if force and decision == "SKIP":
        reasons.append("FORCED")

    # --- Build order command (still DRY-RUN unless caller executes/append --submit) ---
    symbol = build_osi_symbol(idea.ticker, idea.expiry, str(cfg["right"]), idea.strike)

    entry_credit = _round_to_tick(idea.last)
    profit_price = _round_to_tick(entry_credit * float(cfg["profit_take_frac"]))
    stop_price = _round_to_tick(entry_credit * float(cfg["stop_loss_mult"]))

    stop_limit_price = None
    if bool(cfg.get("use_stop_limit", False)):
        stop_limit_price = _round_to_tick(entry_credit * float(cfg.get("stop_limit_mult", cfg["stop_loss_mult"])))

    parts = [
        sys.executable,
        "-m",
        str(cfg["oms_cli_module"]),
        "place-trigger-oco",
    ]
    if account:
        parts += ["--account", str(account)]

    parts += [
        "--symbol", symbol,
        "--qty", str(int(cfg["qty"])),
        "--entry-instruction", "SELL_TO_OPEN",
        "--entry-type", str(cfg["entry_type"]),
        "--entry-price", f"{entry_credit:.2f}",
        "--profit-price", f"{profit_price:.2f}",
        "--stop-price", f"{stop_price:.2f}",
        "--complex-type", str(cfg["complex_type"]),
        "--session", str(cfg["session"]),
        "--duration", str(cfg["duration"]),
    ]
    if stop_limit_price is not None:
        parts += ["--stop-limit-price", f"{stop_limit_price:.2f}"]

    return decision, _join_reasons(reasons), " ".join(parts)


def score_ideas(df: pd.DataFrame, cfg: Dict[str, Any], *, account: Optional[str], force: bool) -> pd.DataFrame:
    out_rows: List[Dict[str, Any]] = []
    for _, r in df.iterrows():
        idea = _extract_idea(r)
        if not idea:
            continue

        decision, reasons, cmd = decide_and_build_command(idea, cfg, account=account, force=force)

        out_rows.append(
            {
                "Ticker": idea.ticker,
                "Expiry": idea.expiry.date().isoformat(),
                "Strike": idea.strike,
                "Underlying": idea.underlying,
                "Last": idea.last,
                "Open Interest": idea.open_interest,
                "DTE": idea.dte,
                "Win Rate (%)": idea.win_rate,
                "Annualised Simple (%)": idea.annualised_roi,
                "Earnings": idea.earnings.date().isoformat() if idea.earnings else "",
                "Decision": decision,
                "Reasons": reasons,
                "OMS Command (DRY-RUN)": cmd,
            }
        )

    odf = pd.DataFrame(out_rows)
    if not odf.empty:
        odf = odf.sort_values(
            by=["Decision", "Annualised Simple (%)", "Win Rate (%)", "Open Interest"],
            ascending=[True, False, False, False],
            na_position="last",
        ).reset_index(drop=True)
    return odf


def _filter_for_shell(scored: pd.DataFrame, include_review: bool, include_skip: bool, force: bool) -> pd.DataFrame:
    if scored.empty:
        return scored
    allowed = {"OK"}
    if include_review:
        allowed.add("REVIEW")
    if include_skip and force:
        allowed.add("SKIP")
    return scored[scored["Decision"].isin(sorted(allowed))].copy()


def write_outputs(scored: pd.DataFrame, *, outdir: Path, basename: str) -> Tuple[Path, Path]:
    outdir.mkdir(parents=True, exist_ok=True)
    xlsx_path = outdir / f"{basename}_ideas_scored.xlsx"
    sh_path = outdir / f"{basename}_oms_commands.sh"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        scored.to_excel(w, index=False, sheet_name="Ideas")

    wb = load_workbook(xlsx_path)
    ws = wb["Ideas"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for col_idx, col_name in enumerate(scored.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in scored[col_name].fillna("").astype(str).head(200)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 65)

    fill_ok = PatternFill("solid", fgColor="C6EFCE")
    fill_review = PatternFill("solid", fgColor="FFEB9C")
    fill_skip = PatternFill("solid", fgColor="FFC7CE")

    col_map = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    dcol = col_map.get("Decision")
    if dcol:
        for r in range(2, ws.max_row + 1):
            v = str(ws.cell(r, dcol).value or "")
            fill = fill_ok if v == "OK" else (fill_review if v == "REVIEW" else fill_skip)
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill

    wb.save(xlsx_path)

    with open(sh_path, "w", encoding="utf-8") as f:
        f.write("#!/usr/bin/env bash\nset -euo pipefail\n\n")
        f.write(f"# Generated: {datetime.now(SGT).isoformat()}\n")
        f.write(f"# Source: {basename}\n\n")
        for _, row in scored.iterrows():
            dec = str(row.get("Decision") or "")
            reasons = str(row.get("Reasons") or "")
            cmd = str(row.get("OMS Command (DRY-RUN)") or "")
            if not cmd:
                continue
            f.write(f"# {dec} {row.get('Ticker')} {row.get('Expiry')} strike={row.get('Strike')} | {reasons}\n")
            f.write(cmd + "\n\n")

    os.chmod(sh_path, 0o755)
    return xlsx_path, sh_path


def run_commands(scored: pd.DataFrame, *, include_review: bool, include_skip: bool, force: bool) -> None:
    run_df = _filter_for_shell(scored, include_review, include_skip, force)
    if run_df.empty:
        print("[execute] nothing to run (after filtering).")
        return

    for i, row in run_df.iterrows():
        cmd = str(row.get("OMS Command (DRY-RUN)") or "")
        if not cmd:
            continue
        cmd2 = cmd + " --submit"
        print(f"[execute] ({i+1}/{len(run_df)}) {row.get('Ticker')} {row.get('Expiry')} | {row.get('Decision')} | {row.get('Reasons')}")
        print(f"[execute] {cmd2}")
        subprocess.run(cmd2, shell=True, check=True)


def _load_config_file(p: Optional[str]) -> Dict[str, Any]:
    if not p:
        return {}
    path = Path(p).expanduser()
    if not path.exists():
        raise SystemExit(f"Config file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _merge_cfg(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(base)
    for k, v in (override or {}).items():
        out[k] = v
    return out


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Score sell-put ideas and generate OMS commands")
    ap.add_argument("--file", required=True, help="Ideas Excel file")
    ap.add_argument("--sheet", default="", help="Sheet name (default: first sheet)")
    ap.add_argument("--outdir", default="", help="Output directory (default: alongside input file)")
    ap.add_argument("--config", default="", help="Optional JSON config to override defaults")
    ap.add_argument("--account", default="", help="Optional: account number or hash (passed to oms_cli)")
    ap.add_argument("--qty", type=int, default=None, help="Override default qty (contracts)")
    ap.add_argument("--include-review", action="store_true", help="Include REVIEW ideas in the generated .sh")
    ap.add_argument("--include-skip", action="store_true", help="Include SKIP ideas (requires --force)")
    ap.add_argument("--force", action="store_true", help="Allow running SKIP ideas (dangerous)")
    ap.add_argument("--execute", action="store_true", help="Actually run the commands (adds --submit)")
    return ap


def main() -> None:
    args = build_parser().parse_args()

    cfg = dict(DEFAULTS)
    if args.config:
        cfg = _merge_cfg(cfg, _load_config_file(args.config))

    if args.qty is not None:
        cfg["qty"] = int(args.qty)

    in_path = Path(args.file).expanduser()
    if not in_path.exists():
        raise SystemExit(f"Input file not found: {in_path}")

    outdir = Path(args.outdir).expanduser() if args.outdir else in_path.parent
    basename = in_path.stem  # traceability

    df = load_ideas_excel(str(in_path), sheet=(args.sheet or None))
    scored = score_ideas(df, cfg, account=(args.account or None), force=bool(args.force))

    stamp = _today_sgt().date().isoformat()
    out_base = f"{stamp}_{basename}"

    xlsx_path, sh_path = write_outputs(scored, outdir=outdir, basename=out_base)

    print(f"[ok] wrote: {xlsx_path}")
    print(f"[ok] wrote: {sh_path}")
    print(
        f"[ok] ideas: total={len(scored)}  "
        f"OK={(scored['Decision']=='OK').sum()}  "
        f"REVIEW={(scored['Decision']=='REVIEW').sum()}  "
        f"SKIP={(scored['Decision']=='SKIP').sum()}"
    )

    if args.execute:
        run_commands(scored, include_review=args.include_review, include_skip=args.include_skip, force=args.force)


if __name__ == "__main__":
    main()
