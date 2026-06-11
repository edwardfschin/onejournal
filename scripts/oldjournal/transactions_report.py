#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/journal/transactions_report.py
Version: 1.3.1
Updated: 2025-12-21 (SGT)

Purpose
-------
Generate a complete journal report from DuckDB canonical reporting views:

Source of truth (views)
-----------------------
- journal.v_transactions_raw_latest        (deduped transaction headers)
- journal.v_transaction_items_norm         (normalized legs/items)
- journal.v_trade_fee_lines_dedup          (fee lines deduped)
- journal.v_transaction_recon_resolved     (final audit view)

Sheets
------
1) Transactions (Schwab-faithful + explicit)
2) Raw (Joined)        : headers + items_norm (NO fee join to avoid cartesian dupes)
3) Fees (Dedup)        : fee lines joined to headers
4) Recon (Resolved)    : audit view
5) By Symbol Activity  : legacy avg-cost engine (best-effort using normalized fields)
6) Summary Dashboard

Notes
-----
- TRADE descriptions: prefer instrument_description from items_norm.
- Non-TRADE descriptions: use header description.
- Transactions sheet is “economic lines”:
    - Non-TRADE: one row per header (Amount = header net_amount)
    - TRADE: principal legs from items_norm (fee_type_norm IS NULL) + fee lines from v_trade_fee_lines_dedup
- Running Balance = cumulative sum of Amount by (Date, Time, Ref #, Line Type).
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

import duckdb
import pandas as pd


# =============================================================================
# CONFIG
# =============================================================================
DEBUG = True

OUTPUT_BASE = (
    Path.home()
    / "Library"
    / "Mobile Documents"
    / "com~apple~CloudDocs"
    / "Projects"
    / "TradersGPS"
    / "output"
    / "journal"
    / "transactions"
)

DB_PATH = Path.home() / "tgps-project" / "data" / "journal" / "tgps_trades.duckdb"

REL_HEADERS = "journal.v_transactions_raw_latest"
REL_ITEMS = "journal.v_transaction_items_norm"
REL_FEE_LINES = "journal.v_trade_fee_lines_dedup"
REL_RECON = "journal.v_transaction_recon_resolved"

# =============================================================================
# HELPERS
# =============================================================================
import json


def _json_obj(x):
    if x is None:
        return None
    if isinstance(x, dict):
        return x
    if isinstance(x, str):
        try:
            return json.loads(x)
        except Exception:
            return None
    return None


def extract_primary_symbol_from_payload(payload) -> tuple[str | None, str | None]:
    """
    Non-TRADE: symbol = first transferItems[*].instrument.symbol (prefer non-CURRENCY).
    Returns (symbol, asset_type).
    """
    obj = _json_obj(payload)
    if not obj:
        return (None, None)

    items = obj.get("transferItems") or []
    if not isinstance(items, list) or not items:
        return (None, None)

    best = None
    for it in items:
        inst = (it or {}).get("instrument") or {}
        sym = inst.get("symbol")
        at = (inst.get("assetType") or "").upper() or None
        if sym and at != "CURRENCY":
            best = (sym, at)
            break

    if best:
        return best

    inst0 = (items[0] or {}).get("instrument") or {}
    return (inst0.get("symbol"), (inst0.get("assetType") or "").upper() or None)


def build_attached_underlying_map(items: pd.DataFrame) -> pd.DataFrame:
    """
    For TRADE fees: attach to the trade's underlying using principal legs only.
    Rule: pick first principal leg preferring non-CURRENCY, ordered by leg_index.
    """
    df = items.copy()

    principal = df[df["fee_type_norm"].isna()].copy()
    if principal.empty:
        return pd.DataFrame(
            columns=[
                "account_hash",
                "activity_id",
                "attached_underlying",
                "multi_underlying",
            ]
        )

    principal["asset_u"] = (
        principal["asset_type_norm"].fillna("").astype(str).str.upper()
    )
    principal["is_currency"] = (principal["asset_u"] == "CURRENCY").astype(int)

    def _attached(row) -> str | None:
        return (
            row.get("instrument_underlying_symbol")
            or row.get("symbol_norm")
            or row.get("instrument_root_symbol")
            or row.get("instrument_symbol")
        )

    principal["attached_underlying"] = principal.apply(_attached, axis=1)

    # multi-underlying flag for audit
    nunq = (
        principal.groupby(["account_hash", "activity_id"])["attached_underlying"]
        .nunique(dropna=True)
        .reset_index()
        .rename(columns={"attached_underlying": "n_underlyings"})
    )
    nunq["multi_underlying"] = (nunq["n_underlyings"] > 1).astype(int)

    principal = principal.sort_values(
        ["account_hash", "activity_id", "is_currency", "leg_index"], kind="mergesort"
    )
    pick = principal.groupby(["account_hash", "activity_id"], as_index=False).head(1)[
        ["account_hash", "activity_id", "attached_underlying"]
    ]

    out = pick.merge(
        nunq[["account_hash", "activity_id", "multi_underlying"]],
        on=["account_hash", "activity_id"],
        how="left",
    )
    out["multi_underlying"] = out["multi_underlying"].fillna(0).astype(int)
    return out


# =============================================================================
# EXCEL SANITIZER (avoid openpyxl Cannot convert <NA>)
# =============================================================================
def df_clean_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df = df.where(pd.notnull(df), None)

    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].astype("string").replace({"NaT": None})

    for col in df.columns:
        cleaned = []
        for val in df[col]:
            if val is None or val is pd.NA:
                cleaned.append(None)
                continue
            if isinstance(val, float) and pd.isna(val):
                cleaned.append(None)
                continue
            if isinstance(val, (dict, list)):
                cleaned.append(str(val))
                continue
            if not isinstance(val, (int, float, str, bool, type(None))):
                cleaned.append(str(val))
                continue
            cleaned.append(val)
        df[col] = cleaned

    return df


# =============================================================================
# REGEX / PARSERS (kept for activity sheet)
# =============================================================================
RE_ASSIGN = re.compile(
    r"ASG:\s*([0-9.]+)\s*\.([A-Z]+)(\d{6})([CP])(\d+(\.\d+)?)",
    re.IGNORECASE,
)
RE_EXP = re.compile(
    r"EXP:\s*([0-9.]+)\s*\.([A-Z]+)(\d{6})([CP])(\d+(\.\d+)?)",
    re.IGNORECASE,
)

RE_UPON = re.compile(r"\bUPON\s+([A-Z]{1,5})\b", re.IGNORECASE)
RE_UPON_QTY = re.compile(r"\bBOT\s+([0-9.]+)\s+([A-Z]{1,5})\s+UPON\b", re.IGNORECASE)

RESERVED_WORDS = {
    "BOT",
    "SLD",
    "SELL",
    "BUY",
    "UPON",
    "DIV",
    "DOI",
    "JRN",
    "BAL",
    "EXP",
    "RAD",
    "ADR",
    "PUT",
    "CALL",
    "INT",
}

RE_TILDE_TICKER = re.compile(r"~\s*([A-Z.\-]{1,10})\s*$")


def detect_assignment_from_desc(desc: str) -> dict | None:
    desc_u = (desc or "").upper()

    m = RE_ASSIGN.search(desc_u)
    if m:
        qty_contracts = float(m.group(1))
        symbol = m.group(2).upper()
        expiry = datetime.strptime(m.group(3), "%y%m%d").date()
        cp = m.group(4).upper()
        strike = float(m.group(5))
        shares = int(qty_contracts * 100)
        side = "BUY" if cp == "P" else "SELL"
        return {
            "symbol": symbol,
            "expiry": expiry,
            "strike": strike,
            "shares": shares,
            "side": side,
            "event": "ASSIGN",
        }

    m2 = RE_UPON_QTY.search(desc_u)
    if m2:
        qty = float(m2.group(1))
        symbol = m2.group(2).upper()
        return {
            "symbol": symbol,
            "expiry": None,
            "strike": None,
            "shares": int(qty),
            "side": "BUY",
            "event": "ASSIGN_UPON",
        }

    m3 = RE_UPON.search(desc_u)
    if m3:
        symbol = m3.group(1).upper()
        return {
            "symbol": symbol,
            "expiry": None,
            "strike": None,
            "shares": 0,
            "side": "BUY",
            "event": "ASSIGN_UPON",
        }

    return None


def detect_expiration_from_desc(desc: str) -> dict | None:
    desc_u = (desc or "").upper()
    m = RE_EXP.search(desc_u)
    if not m:
        return None
    qty_contracts = float(m.group(1))
    symbol = m.group(2).upper()
    expiry = datetime.strptime(m.group(3), "%y%m%d").date()
    strike = float(m.group(5))
    shares = int(qty_contracts * 100)
    return {
        "symbol": symbol,
        "expiry": expiry,
        "strike": strike,
        "shares": shares,
        "event": "EXPIRATION",
    }


def extract_symbol_from_description(desc: str | None) -> str | None:
    if not desc:
        return None
    for tok in re.findall(r"\b([A-Z]{1,5})\b", desc):
        if tok not in RESERVED_WORDS:
            return tok
    return None


# =============================================================================
# LOADERS (hard-wired to your schemas)
# =============================================================================
def load_headers(con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    df = con.execute(
        f"""
        SELECT
            account_hash,
            fetched_at,
            from_iso,
            to_iso,
            activity_payload,
            activity_id,
            order_id,
            time,
            trade_date,
            settlement_date,
            type,
            status,
            sub_account,
            net_amount,
            description
        FROM {REL_HEADERS}
        ORDER BY trade_date, time, activity_id
        """
    ).df()

    # trade_date is DATE in DuckDB -> keep as tz-naive datetime64[ns]
    df["trade_date"] = pd.to_datetime(df["trade_date"], errors="coerce")

    # time is TIMESTAMP WITH TIME ZONE -> keep tz-aware (UTC) for sorting/display
    # (Excel will still be fine because we stringify datetimes in df_clean_for_excel)
    df["time"] = pd.to_datetime(df["time"], errors="coerce", utc=True)

    # Fill missing trade_date from time, but IMPORTANT:
    # convert to Asia/Singapore then drop tz so trade_date stays tz-naive dtype.
    mask = df["trade_date"].isna() & df["time"].notna()
    if mask.any():
        df.loc[mask, "trade_date"] = (
            df.loc[mask, "time"]
            .dt.tz_convert("Asia/Singapore")
            .dt.tz_localize(None)
            .dt.normalize()
        )
    # Extract a primary symbol for non-TRADE rows from activity_payload (deterministic)
    sym_asset = df["activity_payload"].apply(extract_primary_symbol_from_payload)
    df["header_symbol_first"] = sym_asset.apply(lambda t: t[0])
    df["header_asset_type_first"] = sym_asset.apply(lambda t: t[1])

    return df


def load_items(con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    df = con.execute(
        f"""
        SELECT
            account_hash,
            activity_id,
            leg_index,

            instrument_asset_type,
            instrument_type,
            instrument_symbol,
            instrument_underlying_symbol,
            instrument_root_symbol,
            instrument_description,

            symbol_norm,
            asset_type_norm,
            position_effect_norm,
            put_call_norm,
            strike_norm,
            expiry_norm,
            price_norm,
            amount_norm,
            fee_type_norm,
            cost_norm,

            extras
        FROM {REL_ITEMS}
        ORDER BY activity_id, leg_index
        """
    ).df()

    df["expiry_norm"] = pd.to_datetime(df["expiry_norm"], errors="coerce").dt.date
    for c in ("strike_norm", "price_norm", "amount_norm", "cost_norm"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def load_fee_lines(con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    df = con.execute(
        f"""
        SELECT
            account_hash,
            activity_id,
            fee_type,
            fee_cost,
            asset_type
        FROM {REL_FEE_LINES}
        ORDER BY activity_id, fee_type
        """
    ).df()

    df["fee_cost"] = pd.to_numeric(df["fee_cost"], errors="coerce")
    return df


def load_recon(con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    # You have this view; load everything (diagnostic sheet)
    return con.execute(
        f"SELECT * FROM {REL_RECON} ORDER BY trade_date, time, activity_id"
    ).df()


# =============================================================================
# HELPERS: Underlying parsing
# =============================================================================
def underlying_from_header_desc(desc: str | None) -> str | None:
    if not desc:
        return None
    m = RE_TILDE_TICKER.search(desc.strip().upper())
    if m:
        return m.group(1)
    # fallback: last ticker-ish token
    m2 = re.search(r"\b([A-Z]{1,6})\b\s*$", desc.strip().upper())
    return m2.group(1) if m2 else None


# =============================================================================
# BUILD: Transactions sheet
# =============================================================================
def build_transactions_sheet(
    hdr: pd.DataFrame,
    items: pd.DataFrame,
    fees: pd.DataFrame,
) -> pd.DataFrame:
    """
    Transactions sheet (economic lines)

    - Non-TRADE: one row per header
        Amount = header net_amount
        Symbol/Underlying = header_symbol_first (from activity_payload transferItems), else blank

    - TRADE: principal legs from items_norm where fee_type_norm IS NULL
        Symbol = instrument_symbol (fallback symbol_norm)
        Underlying = instrument_underlying_symbol / root / symbol_norm (best effort)

    - Fees (from v_trade_fee_lines_dedup):
        Symbol/Underlying = attached_underlying via activity_id, else blank (NO parsing)

    - Currency principal legs (asset_type_norm == CURRENCY and fee_type_norm IS NULL):
        Symbol/Underlying = attached_underlying via activity_id, else blank (NO parsing)

    IMPORTANT:
    - If attached_underlying cannot be resolved from activity_id, leave blank as requested.
    """
    hdr = hdr.copy()
    items = items.copy()
    fees = fees.copy()

    # Map activity_id -> attached_underlying based on principal (non-fee) legs
    attached = build_attached_underlying_map(items)

    # Join headers into items for TRADE context
    it = items.merge(
        hdr[
            [
                "account_hash",
                "activity_id",
                "trade_date",
                "time",
                "type",
                "description",
                "net_amount",
                "order_id",
                "settlement_date",
                "status",
                "sub_account",
            ]
        ],
        on=["account_hash", "activity_id"],
        how="left",
    )

    # ----------------------------
    # Non-TRADE headers: 1 row each
    # ----------------------------
    non_trade = hdr[hdr["type"].fillna("").str.upper() != "TRADE"].copy()

    non_trade_rows = pd.DataFrame(
        {
            "Date": pd.to_datetime(non_trade["trade_date"], errors="coerce").dt.date,
            "Time": pd.to_datetime(non_trade["time"], errors="coerce"),
            "Type (Schwab)": non_trade["type"],
            "Instrument Type": non_trade.get("header_asset_type_first"),
            "Economic Role": non_trade["type"],
            "Description": non_trade["description"],
            # Use payload-derived primary symbol; if None, leave blank
            "Symbol": non_trade.get("header_symbol_first"),
            "Underlying": non_trade.get("header_symbol_first"),
            "Qty": None,
            "Price": None,
            "Amount": pd.to_numeric(non_trade["net_amount"], errors="coerce").fillna(
                0.0
            ),
            "Ref #": non_trade["activity_id"],
        }
    )

    # ----------------------------
    # Principal legs (fee_type_norm IS NULL) split into:
    #   A) non-currency principal legs (normal TRADE_LEG lines)
    #   B) currency principal legs (attach via activity_id)
    # ----------------------------
    trade = it[it["type"].fillna("").str.upper() == "TRADE"].copy()
    principal_all = trade[trade["fee_type_norm"].isna()].copy()

    principal_all["asset_u"] = (
        principal_all["asset_type_norm"].fillna("").astype(str).str.upper()
    )
    principal_currency = principal_all[principal_all["asset_u"] == "CURRENCY"].copy()
    principal_noncur = principal_all[principal_all["asset_u"] != "CURRENCY"].copy()

    # ---- A) Non-currency principal legs (TRADE_LEG) ----
    def _symbol(row) -> str | None:
        return row.get("instrument_symbol") or row.get("symbol_norm")

    def _instrument_type(row) -> str | None:
        return (
            row.get("instrument_type")
            or row.get("asset_type_norm")
            or row.get("instrument_asset_type")
        )

    def _underlying(row) -> str | None:
        u = row.get("instrument_underlying_symbol")
        if u and str(u).strip():
            return u
        root = row.get("instrument_root_symbol")
        if root and str(root).strip():
            return root
        sym_norm = row.get("symbol_norm")
        if sym_norm and str(sym_norm).strip():
            return sym_norm
        # last resort: if nothing, leave None (don’t invent)
        return None

    def _desc(row) -> str | None:
        d = row.get("instrument_description")
        if d and str(d).strip():
            return d
        return row.get("description")

    # Amount: prefer cost_norm, fallback to qty*price (option multiplier if put_call_norm present)
    def _amount(row) -> float:
        c = row.get("cost_norm")
        if c is not None and pd.notna(c):
            return float(c)

        qty = row.get("amount_norm")
        px = row.get("price_norm")
        if qty is None or px is None or pd.isna(qty) or pd.isna(px):
            return 0.0

        pc = row.get("put_call_norm")
        if pc is not None and str(pc).strip():
            return float(qty) * float(px) * 100.0
        return float(qty) * float(px)

    trade_leg_rows = pd.DataFrame(
        {
            "Date": pd.to_datetime(
                principal_noncur["trade_date"], errors="coerce"
            ).dt.date,
            "Time": pd.to_datetime(principal_noncur["time"], errors="coerce"),
            "Type (Schwab)": principal_noncur["type"],
            "Instrument Type": principal_noncur.apply(_instrument_type, axis=1),
            "Economic Role": "TRADE_LEG",
            "Description": principal_noncur.apply(_desc, axis=1),
            "Symbol": principal_noncur.apply(_symbol, axis=1),
            "Underlying": principal_noncur.apply(_underlying, axis=1),
            "Qty": pd.to_numeric(principal_noncur["amount_norm"], errors="coerce"),
            "Price": pd.to_numeric(principal_noncur["price_norm"], errors="coerce"),
            "Amount": principal_noncur.apply(_amount, axis=1),
            "Ref #": principal_noncur["activity_id"],
        }
    )

    # Defensive de-dupe: keep the “better described” row if duplicates exist
    if not trade_leg_rows.empty:
        trade_leg_rows["_desc_ok"] = (
            trade_leg_rows["Description"].fillna("").astype(str).str.len()
        )
        trade_leg_rows["_und_ok"] = (
            trade_leg_rows["Underlying"].fillna("").astype(str).str.len()
        )
        trade_leg_rows = trade_leg_rows.sort_values(
            ["Ref #", "Symbol", "_desc_ok", "_und_ok"],
            ascending=[True, True, False, False],
            kind="mergesort",
        )
        trade_leg_rows = trade_leg_rows.drop_duplicates(
            subset=["Ref #", "Symbol", "Qty", "Price", "Amount"],
            keep="first",
        ).drop(columns=["_desc_ok", "_und_ok"])

    # ---- B) Currency principal legs (attach via activity_id; else blank) ----
    # Join attached underlying
    cur = principal_currency.merge(
        attached[["account_hash", "activity_id", "attached_underlying"]],
        on=["account_hash", "activity_id"],
        how="left",
    )

    currency_rows = pd.DataFrame(
        {
            "Date": pd.to_datetime(cur["trade_date"], errors="coerce").dt.date,
            "Time": pd.to_datetime(cur["time"], errors="coerce"),
            "Type (Schwab)": cur["type"],
            "Instrument Type": cur["asset_type_norm"].where(
                cur["asset_type_norm"].notna(), "CURRENCY"
            ),
            "Economic Role": "CURRENCY",
            "Description": cur["instrument_description"].where(
                cur["instrument_description"].notna(), cur["description"]
            ),
            # As requested: ONLY attach via activity_id; if not found -> blank
            "Symbol": cur["attached_underlying"],
            "Underlying": cur["attached_underlying"],
            "Qty": pd.to_numeric(cur["amount_norm"], errors="coerce"),
            "Price": pd.to_numeric(cur["price_norm"], errors="coerce"),
            "Amount": cur.apply(_amount, axis=1),
            "Ref #": cur["activity_id"],
        }
    )

    # ----------------------------
    # Fees (dedup view): attach via activity_id; else blank
    # ----------------------------
    fee_rows = fees.merge(
        hdr[["account_hash", "activity_id", "trade_date", "time", "type"]],
        on=["account_hash", "activity_id"],
        how="left",
    ).merge(
        attached[["account_hash", "activity_id", "attached_underlying"]],
        on=["account_hash", "activity_id"],
        how="left",
    )

    fee_out = pd.DataFrame(
        {
            "Date": pd.to_datetime(fee_rows["trade_date"], errors="coerce").dt.date,
            "Time": pd.to_datetime(fee_rows["time"], errors="coerce"),
            "Type (Schwab)": fee_rows["type"].where(fee_rows["type"].notna(), "TRADE"),
            "Instrument Type": fee_rows["asset_type"].where(
                fee_rows["asset_type"].notna(), "FEE"
            ),
            "Economic Role": fee_rows["fee_type"],
            "Description": fee_rows["fee_type"],
            # As requested: ONLY attach via activity_id; if not found -> blank
            "Symbol": fee_rows["attached_underlying"],
            "Underlying": fee_rows["attached_underlying"],
            "Qty": None,
            "Price": None,
            "Amount": pd.to_numeric(fee_rows["fee_cost"], errors="coerce").fillna(0.0),
            "Ref #": fee_rows["activity_id"],
        }
    )

    # ----------------------------
    # Combine + running balance
    # ----------------------------
    tx = pd.concat(
        [non_trade_rows, trade_leg_rows, currency_rows, fee_out], ignore_index=True
    )

    # Stable sort: Date, Time, Ref #, then push currency/fees after legs within same activity
    tx["_sort_date"] = pd.to_datetime(tx["Date"], errors="coerce")
    tx["_sort_time"] = pd.to_datetime(tx["Time"], errors="coerce", utc=True)
    tx["_sort_ref"] = pd.to_numeric(tx["Ref #"], errors="coerce")

    def _after_legs(role) -> int:
        s = "" if role is None else str(role)
        if s == "CURRENCY":
            return 1
        if s.endswith("_FEE") or s in {
            "COMMISSION",
            "SEC_FEE",
            "TAF_FEE",
            "GST_FEE",
            "OPT_REG_FEE",
        }:
            return 2
        return 0

    tx["_sort_tail"] = tx["Economic Role"].apply(_after_legs)

    tx = (
        tx.sort_values(
            ["_sort_date", "_sort_time", "_sort_ref", "_sort_tail"], kind="mergesort"
        )
        .drop(columns=["_sort_date", "_sort_time", "_sort_ref", "_sort_tail"])
        .reset_index(drop=True)
    )

    tx["Running Balance"] = (
        pd.to_numeric(tx["Amount"], errors="coerce").fillna(0.0).cumsum()
    )

    cols = [
        "Date",
        "Time",
        "Type (Schwab)",
        "Instrument Type",
        "Economic Role",
        "Description",
        "Symbol",
        "Underlying",
        "Qty",
        "Price",
        "Amount",
        "Running Balance",
        "Ref #",
    ]
    return tx[cols]


# =============================================================================
# BUILD: Raw joined (headers + items) / Fees / Recon
# =============================================================================
def build_raw_joined(hdr: pd.DataFrame, items: pd.DataFrame) -> pd.DataFrame:
    df = items.merge(
        hdr,
        on=["account_hash", "activity_id"],
        how="left",
        suffixes=("_item", "_hdr"),
    )

    # Put header columns first for readability
    header_first = [
        "account_hash",
        "activity_id",
        "order_id",
        "time",
        "trade_date",
        "settlement_date",
        "type",
        "status",
        "sub_account",
        "net_amount",
        "description",
        "fetched_at",
        "from_iso",
        "to_iso",
    ]
    cols = list(df.columns)
    ordered = [c for c in header_first if c in cols] + [
        c for c in cols if c not in header_first
    ]
    return df[ordered]


def build_fees_sheet(hdr: pd.DataFrame, fees: pd.DataFrame) -> pd.DataFrame:
    df = fees.merge(
        hdr[
            [
                "account_hash",
                "activity_id",
                "time",
                "trade_date",
                "type",
                "description",
                "net_amount",
            ]
        ],
        on=["account_hash", "activity_id"],
        how="left",
    )
    df["fee_cost"] = pd.to_numeric(df["fee_cost"], errors="coerce")
    return df


# =============================================================================
# BY SYMBOL ACTIVITY + AVG COST ENGINE (best-effort)
# =============================================================================
def build_symbol_activity(
    raw_joined: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = raw_joined.copy()

    activity_rows: list[dict] = []

    for _, row in df.iterrows():
        symbol = row.get("symbol_norm") or row.get("instrument_symbol")
        if not symbol:
            symbol = extract_symbol_from_description(row.get("description"))

        pos_eff = (row.get("position_effect_norm") or "").upper()
        qty = 0.0
        if row.get("amount_norm") is not None:
            try:
                q = float(row.get("amount_norm") or 0.0)
            except Exception:
                q = 0.0

            # OPENING/CLOSING sign inference (legacy)
            if pos_eff == "OPENING":
                qty = abs(q)
            elif pos_eff == "CLOSING":
                qty = -abs(q)
            else:
                qty = q

        # Prefer leg cash if present; fallback header net_amount
        try:
            net = (
                float(row.get("cost_norm"))
                if row.get("cost_norm") is not None
                else float(row.get("net_amount") or 0.0)
            )
        except Exception:
            net = 0.0

        base_event = {
            "symbol": symbol,
            "date": row.get("trade_date"),
            "time": row.get("time"),
            "type": row.get("type"),
            "side": pos_eff,
            "qty": qty,
            "price": row.get("price_norm"),
            "net": net,
            "description": row.get("description") or "",
        }
        activity_rows.append(base_event)

        desc = row.get("description") or ""
        assign = detect_assignment_from_desc(desc)
        if assign:
            shares = assign["shares"]
            strike = float(assign["strike"] or 0.0) if assign["strike"] else 0.0
            if assign["side"] == "BUY":
                qty_shares = shares
                net_shares = -shares * strike
            else:
                qty_shares = -shares
                net_shares = shares * strike

            activity_rows.append(
                {
                    "symbol": assign["symbol"],
                    "date": row.get("trade_date"),
                    "time": row.get("time"),
                    "type": "ASSIGN",
                    "side": assign["side"],
                    "qty": qty_shares,
                    "price": strike if strike else None,
                    "net": net_shares,
                    "description": desc,
                }
            )

        exp = detect_expiration_from_desc(desc)
        if exp:
            activity_rows.append(
                {
                    "symbol": exp["symbol"],
                    "date": row.get("trade_date"),
                    "time": row.get("time"),
                    "type": "EXPIRATION",
                    "side": "",
                    "qty": 0.0,
                    "price": None,
                    "net": net,
                    "description": desc,
                }
            )

    if not activity_rows:
        return pd.DataFrame(), pd.DataFrame()

    df_act = pd.DataFrame(activity_rows)
    df_act["date"] = pd.to_datetime(df_act["date"], errors="coerce")
    df_act["time"] = pd.to_datetime(df_act["time"], errors="coerce")
    df_act = df_act.sort_values(["symbol", "date", "time"])

    expanded_rows: list[dict] = []
    summary: dict[str, dict] = {}

    for symbol, grp in df_act.groupby("symbol"):
        if symbol is None:
            continue

        avg_cost = 0.0
        qty_pos = 0.0
        realised_pl = 0.0
        div_total = 0.0
        int_total = 0.0
        fee_total = 0.0

        for _, r in grp.iterrows():
            row_dict = dict(r)

            r_type = (r.get("type") or "").upper()
            side = (r.get("side") or "").upper()
            qty = float(r.get("qty") or 0.0)
            net = float(r.get("net") or 0.0)
            desc = (r.get("description") or "").upper()

            if r_type in {"TRADE", "ASSIGN"}:
                if side == "OPENING" and qty > 0:
                    if r.get("price") is not None:
                        total_cost = float(r["price"]) * qty
                        new_qty = qty_pos + qty
                        if new_qty > 0:
                            avg_cost = ((avg_cost * qty_pos) + total_cost) / new_qty
                        qty_pos = new_qty

                elif side == "CLOSING" and qty < 0:
                    close_qty = abs(qty)
                    sale_value = float(r.get("price") or 0.0) * close_qty
                    realised_pl += sale_value - (avg_cost * close_qty)
                    qty_pos -= close_qty
                    if qty_pos <= 0:
                        qty_pos = 0.0
                        avg_cost = 0.0

            if "DIVIDEND" in desc or r_type == "DIVIDEND_OR_INTEREST":
                div_total += net
            if "INTEREST" in desc:
                int_total += net
            if r_type in {"JOURNAL", "SMA_ADJUSTMENT"} or "ADR" in desc:
                fee_total += net

            row_dict["avg_cost"] = avg_cost
            row_dict["remaining_qty"] = qty_pos
            expanded_rows.append(row_dict)

        summary[symbol] = {
            "realised": realised_pl,
            "div": div_total,
            "int": int_total,
            "fee": fee_total,
            "net_pl": realised_pl + div_total + int_total + fee_total,
            "remaining_qty": qty_pos,
            "avg_cost": avg_cost,
        }

    df_records = pd.DataFrame(expanded_rows)
    df_summary = pd.DataFrame.from_dict(summary, orient="index")
    df_summary.index.name = "symbol"
    return df_records, df_summary


def build_dashboard(summary_df: pd.DataFrame) -> pd.DataFrame:
    df = summary_df.reset_index()
    if "net_pl" in df.columns:
        df = df.sort_values("net_pl", ascending=False)
    return df


# =============================================================================
# SAVE EXCEL
# =============================================================================
def save_excel(
    transactions: pd.DataFrame,
    raw_joined: pd.DataFrame,
    fees_sheet: pd.DataFrame,
    recon_sheet: pd.DataFrame,
    activity: pd.DataFrame,
    dashboard: pd.DataFrame,
    outpath: Path,
) -> None:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    def append_df(ws, df: pd.DataFrame):
        for r in dataframe_to_rows(df, index=False, header=True):
            clean_row = []
            for x in r:
                if x is pd.NA or (isinstance(x, float) and pd.isna(x)):
                    clean_row.append(None)
                else:
                    clean_row.append(x)
            ws.append(clean_row)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

    def autosize(ws, max_cols: int = 80, max_width: int = 65):
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            if col_idx > max_cols:
                break
            max_len = 0
            for c in list(col_cells)[:300]:
                v = c.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            width = min(max(10, max_len + 2), max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    outpath.parent.mkdir(parents=True, exist_ok=True)

    transactions = df_clean_for_excel(transactions)
    raw_joined = df_clean_for_excel(raw_joined)
    fees_sheet = df_clean_for_excel(fees_sheet)
    recon_sheet = df_clean_for_excel(recon_sheet)
    activity = df_clean_for_excel(activity)
    dashboard = df_clean_for_excel(dashboard)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Transactions"
    append_df(ws, transactions)
    autosize(ws)

    ws2 = wb.create_sheet("Raw (Joined)")
    append_df(ws2, raw_joined)
    autosize(ws2)

    ws3 = wb.create_sheet("Fees (Dedup)")
    append_df(ws3, fees_sheet)
    autosize(ws3)

    ws4 = wb.create_sheet("Recon (Resolved)")
    append_df(ws4, recon_sheet)
    autosize(ws4)

    ws5 = wb.create_sheet("By Symbol Activity")
    if not activity.empty:
        append_df(ws5, activity)
        autosize(ws5)
    else:
        ws5.append(["NO ACTIVITY"])

    ws6 = wb.create_sheet("Summary Dashboard")
    if not dashboard.empty:
        append_df(ws6, dashboard)
        autosize(ws6)
    else:
        ws6.append(["NO SUMMARY"])

    wb.save(outpath)


# =============================================================================
# MAIN
# =============================================================================
def main() -> None:
    today_str = datetime.now().strftime("%Y-%m-%d")
    outpath = OUTPUT_BASE / f"{today_str}_transactions_report.xlsx"

    print(f"[transactions_report] Using DB: {DB_PATH}")

    con = duckdb.connect(str(DB_PATH))

    # Basic existence checks
    for rel in (REL_HEADERS, REL_ITEMS, REL_FEE_LINES, REL_RECON):
        try:
            con.execute(f"SELECT 1 FROM {rel} LIMIT 1")
        except Exception as e:
            con.close()
            raise SystemExit(f"Missing or unreadable relation: {rel}\n{e}")

    hdr = load_headers(con)
    items = load_items(con)
    fees = load_fee_lines(con)
    recon = load_recon(con)

    con.close()

    if DEBUG:
        print(f"[DEBUG] headers rows: {len(hdr):,}")
        print(f"[DEBUG] items rows:   {len(items):,}")
        print(f"[DEBUG] fees rows:    {len(fees):,}")
        print(f"[DEBUG] recon rows:   {len(recon):,}")

    print("[transactions_report] Building sheets…")

    tx_sheet = build_transactions_sheet(hdr, items, fees)
    raw_joined = build_raw_joined(hdr, items)
    fees_sheet = build_fees_sheet(hdr, fees)

    activity_sheet, summary_sheet = build_symbol_activity(raw_joined)
    dashboard_sheet = (
        build_dashboard(summary_sheet) if not summary_sheet.empty else pd.DataFrame()
    )

    save_excel(
        transactions=tx_sheet,
        raw_joined=raw_joined,
        fees_sheet=fees_sheet,
        recon_sheet=recon,
        activity=activity_sheet,
        dashboard=dashboard_sheet,
        outpath=outpath,
    )

    print(f"[transactions_report] Report generated: {outpath}")


if __name__ == "__main__":
    main()
