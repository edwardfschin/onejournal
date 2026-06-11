#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/journal/report_open_orders_live.py
Version: 0.6.2
Updated: 2025-12-24 (SGT)

Purpose
-------
Generate a troubleshooting + audit Excel report from DuckDB:

- journal.open_orders_live        (current actionable open-orders legs; overwritten per run)
- journal.open_orders_snapshots   (append-only polling history of open orders)
- journal.orders_raw              (raw Schwab payload per API call; source of truth)

Sheets
------
1) Summary
2) Dashboard                 (WORKING + PENDING_ACTIVATION only; root-group aware; user-first)
3) Raw                       (latest orders_raw payload flattened to per-leg + containers)
4) Live_Legs_All
5) Live_Open_Only
6) Live_Actionable_Legs
7) Grouped_Orders
8) Snapshots_Legs
9) Snapshot_Changes
10) Orders_Raw_Meta
"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple

import duckdb
import pandas as pd

# ----------------------------
# Config
# ----------------------------

OPEN_STATUSES = {"WORKING", "PENDING_ACTIVATION", "QUEUED"}
DASHBOARD_STATUSES = {"WORKING", "PENDING_ACTIVATION"}

REPORT_LOOKBACK_DAYS = int(os.environ.get("TGPS_REPORT_LOOKBACK_DAYS", "60"))
MAX_RAW_JSON_CELL_CHARS = int(os.environ.get("TGPS_REPORT_MAX_JSON_CHARS", "25000"))

DASHBOARD_COLUMNS = [
    "Order ID",
    "Parent Order ID",
    "Root Order ID",
    "Placed Date",
    "Strategy",
    "Complex Strategy",
    "Instruction",
    "Quantity",
    "Underlying",
    "Expiry",
    "Strike",
    "Order Type",
    "Put/Call",
    "Limit Price",
    "Stop Price",
    "Duration",
    "Mark",
    "Status",
]


def get_db_path() -> str:
    default = os.path.expanduser("~/tgps-project/data/journal/tgps_trades.duckdb")
    return os.environ.get("TGPS_JOURNAL_DB", default)


def out_dir() -> Path:
    return Path(os.path.expanduser("~/tgps-project/output/journal/orders"))


# ----------------------------
# Small helpers
# ----------------------------

def _to_int64_nullable(s: pd.Series) -> pd.Series:
    """Pandas nullable int (preserves nulls; avoids floats)."""
    if s is None:
        return s
    return pd.to_numeric(s, errors="coerce").astype("Int64")


def _safe_join(values: Iterable[Any]) -> str:
    vals: List[str] = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        vals.append(s)
    return ", ".join(sorted(set(vals))) if vals else ""


def _bool_any(values: pd.Series) -> bool:
    if values is None or values.empty:
        return False
    return bool(values.fillna(False).astype(bool).any())


def _is_container_row(df: pd.DataFrame) -> pd.Series:
    # Container rows have symbol/instruction/quantity NULL (by design).
    sym = df.get("symbol")
    instr = df.get("instruction")
    qty = df.get("quantity")
    if sym is None or instr is None or qty is None:
        return pd.Series([False] * len(df), index=df.index)
    return sym.isna() & instr.isna() & qty.isna()


def _truncate(s: str | None, n: int) -> str | None:
    if s is None:
        return None
    if len(s) <= n:
        return s
    return s[: n - 3] + "..."


def _now_utc_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def _fmt_yyyy_mm_dd(x: Any) -> Any:
    """Return YYYY-MM-DD (no time). Keeps None as None."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    try:
        t = pd.to_datetime(x, errors="coerce")
        if pd.isna(t):
            return x
        return t.strftime("%Y-%m-%d")
    except Exception:
        return x


def _parse_occ_symbol_for_expiry_strike(
    sym: Any,
) -> Tuple[Optional[str], Optional[float], Optional[str], Optional[str]]:
    """
    Parse OCC option symbol like:
      'AAPL  240119C00180000' or 'HIMS  231117C00008000'
    Returns: (expiry_yyyy_mm_dd, strike_float, put_call, underlying)
    """
    if sym is None or (isinstance(sym, float) and pd.isna(sym)):
        return None, None, None, None
    s = str(sym).strip()
    if not s:
        return None, None, None, None

    s2 = re.sub(r"\s+", "", s).upper()
    m = re.match(r"^([A-Z]{1,6})(\d{6})([CP])(\d{8})$", s2)
    if not m:
        return None, None, None, None

    und, yymmdd, pc, strike8 = m.groups()
    yy = int(yymmdd[0:2])
    mm = int(yymmdd[2:4])
    dd = int(yymmdd[4:6])

    yyyy = 2000 + yy
    expiry = f"{yyyy:04d}-{mm:02d}-{dd:02d}"
    strike = float(int(strike8)) / 1000.0

    return expiry, strike, ("CALL" if pc == "C" else "PUT"), und


# ----------------------------
# Read tables
# ----------------------------

def read_open_orders_live(con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    return con.execute(
        """
        SELECT *
        FROM journal.open_orders_live
        ORDER BY account_hash, entered_time DESC NULLS LAST, order_id, leg_index;
        """
    ).df()


def read_open_orders_snapshots(
    con: duckdb.DuckDBPyConnection, days: int
) -> pd.DataFrame:
    return con.execute(
        f"""
        SELECT *
        FROM journal.open_orders_snapshots
        WHERE snapshot_ts >= now() - INTERVAL '{int(days)} days'
        ORDER BY snapshot_ts DESC, account_hash, order_id, leg_index;
        """
    ).df()


def read_latest_orders_raw_meta(con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    return con.execute(
        """
        SELECT
            fetched_at,
            account_hash,
            from_iso,
            to_iso,
            status_filter,
            http_status,
            payload_sha256,
            source
        FROM journal.orders_raw
        ORDER BY fetched_at DESC
        LIMIT 25;
        """
    ).df()


def read_latest_orders_raw_payload(
    con: duckdb.DuckDBPyConnection,
) -> Tuple[Dict[str, Any] | None, Any]:
    """
    Returns:
      (meta_row_dict, payload_python)
    """
    row = con.execute(
        """
        SELECT
          fetched_at,
          account_hash,
          from_iso,
          to_iso,
          status_filter,
          http_status,
          payload_sha256,
          source,
          CAST(payload AS VARCHAR) AS payload_json
        FROM journal.orders_raw
        ORDER BY fetched_at DESC
        LIMIT 1;
        """
    ).fetchone()

    if not row:
        return None, []

    cols = [
        "fetched_at",
        "account_hash",
        "from_iso",
        "to_iso",
        "status_filter",
        "http_status",
        "payload_sha256",
        "source",
        "payload_json",
    ]
    d = dict(zip(cols, row))
    payload_json = d.pop("payload_json", None)

    try:
        payload = json.loads(payload_json) if payload_json else []
    except Exception:
        payload = []

    return d, payload


# ----------------------------
# RAW flattening (latest payload) for audit sheet
# ----------------------------

def iter_order_tree_report(
    orders: Iterable[Dict[str, Any]],
) -> Iterator[Tuple[Dict[str, Any], Optional[int], Optional[int], int]]:
    def walk(
        ord_obj: Dict[str, Any],
        parent_id: Optional[int],
        root_id: Optional[int],
        depth: int,
    ):
        oid = ord_obj.get("orderId")
        try:
            oid_int = int(oid) if oid is not None else None
        except Exception:
            oid_int = None

        if root_id is None:
            root_id = oid_int

        yield ord_obj, parent_id, root_id, depth

        for ch in ord_obj.get("childOrderStrategies") or []:
            if isinstance(ch, dict):
                yield from walk(ch, oid_int, root_id, depth + 1)

    for top in orders:
        if isinstance(top, dict):
            yield from walk(top, None, None, 0)


def flatten_latest_raw_payload(meta: Dict[str, Any], payload: Any) -> pd.DataFrame:
    """
    Turn latest orders_raw.payload (top-level orders array) into one row per leg.
    Includes ALL statuses (audit/troubleshooting, not just open).
    Lineage comes from the nesting (childOrderStrategies).
    """
    if not payload or not isinstance(payload, list):
        return pd.DataFrame()

    rows: List[Dict[str, Any]] = []

    for ord_obj, parent_id, root_id, depth in iter_order_tree_report(payload):
        order_id = ord_obj.get("orderId")
        legs = ord_obj.get("orderLegCollection") or []

        common = {
            "raw_fetched_at": meta.get("fetched_at"),
            "raw_payload_sha256": meta.get("payload_sha256"),
            "raw_from_iso": meta.get("from_iso"),
            "raw_to_iso": meta.get("to_iso"),
            "account_hash": meta.get("account_hash"),
            "order_id": order_id,
            "parent_order_id": parent_id,
            "root_order_id": root_id,
            "depth": depth,
            "status": ord_obj.get("status"),
            "enteredTime": ord_obj.get("enteredTime"),
            "closeTime": ord_obj.get("closeTime"),
            "orderType": ord_obj.get("orderType"),
            "orderStrategyType": ord_obj.get("orderStrategyType"),
            "complexOrderStrategyType": ord_obj.get("complexOrderStrategyType"),
            "duration": ord_obj.get("duration"),
            "session": ord_obj.get("session"),
            "price": ord_obj.get("price"),
            "stopPrice": ord_obj.get("stopPrice"),
            "filledQuantity": ord_obj.get("filledQuantity"),
            "remainingQuantity": ord_obj.get("remainingQuantity"),
            "cancelable": ord_obj.get("cancelable"),
            "editable": ord_obj.get("editable"),
            "tag": ord_obj.get("tag"),
            "clientOrderId": ord_obj.get("clientOrderId"),
            "orderTag": ord_obj.get("orderTag"),
        }

        # container (no legs)
        if not legs:
            r = dict(common)
            r.update(
                {
                    "leg_index": -1,
                    "legId": None,
                    "instrumentId": None,
                    "instruction": None,
                    "quantity": None,
                    "assetType": None,
                    "symbol": None,
                    "underlyingSymbol": None,
                    "expirationDate": None,
                    "strikePrice": None,
                    "putCall": None,
                    "row_kind": "CONTAINER",
                    "order_json_trunc": _truncate(
                        json.dumps(ord_obj, ensure_ascii=False),
                        MAX_RAW_JSON_CELL_CHARS,
                    ),
                }
            )
            rows.append(r)
            continue

        for i, leg in enumerate(legs):
            inst = (leg.get("instrument") or {}) if isinstance(leg, dict) else {}
            r = dict(common)
            r.update(
                {
                    "leg_index": i,
                    "legId": leg.get("legId") if isinstance(leg, dict) else None,
                    "instrumentId": inst.get("instrumentId"),
                    "instruction": (
                        leg.get("instruction") if isinstance(leg, dict) else None
                    ),
                    "quantity": leg.get("quantity") if isinstance(leg, dict) else None,
                    "assetType": inst.get("assetType") or inst.get("asset_type"),
                    "symbol": inst.get("symbol"),
                    "underlyingSymbol": inst.get("underlyingSymbol"),
                    "expirationDate": inst.get("expirationDate"),
                    "strikePrice": inst.get("strikePrice") or inst.get("strike"),
                    "putCall": inst.get("putCall"),
                    "row_kind": "LEG",
                    "order_json_trunc": _truncate(
                        json.dumps(ord_obj, ensure_ascii=False),
                        MAX_RAW_JSON_CELL_CHARS,
                    ),
                }
            )
            rows.append(r)

    df = pd.DataFrame(rows)
    for c in ["enteredTime", "closeTime", "raw_fetched_at"]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    return df


# ----------------------------
# Live derived + grouped
# ----------------------------

def build_live_derived(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df

    df = df.copy()

    for c in ["entered_time", "close_time", "updated_at"]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    df["is_container"] = _is_container_row(df)
    df["row_kind"] = df["is_container"].map(lambda x: "CONTAINER" if bool(x) else "LEG")

    df["is_open_status"] = (
        df["status"].isin(OPEN_STATUSES) if "status" in df.columns else False
    )

    if "cancelable" in df.columns and "editable" in df.columns:
        df["is_actionable_leg"] = (
            (df["row_kind"] == "LEG")
            & df["is_open_status"]
            & (df["cancelable"].fillna(False) | df["editable"].fillna(False))
        )
    else:
        df["is_actionable_leg"] = False

    return df


def build_grouped_orders(df_live: pd.DataFrame) -> pd.DataFrame:
    if df_live is None or df_live.empty:
        return pd.DataFrame()

    df = df_live.copy()
    if "row_kind" not in df.columns:
        df = build_live_derived(df)

    keys = ["account_hash", "order_id"]

    agg = {
        "account_id": (
            ("account_id", "first")
            if "account_id" in df.columns
            else ("account_hash", "first")
        ),
        "entered_time_min": (
            ("entered_time", "min")
            if "entered_time" in df.columns
            else ("order_id", "min")
        ),
        "updated_at_max": (
            ("updated_at", "max") if "updated_at" in df.columns else ("order_id", "max")
        ),
        "statuses": (
            ("status", _safe_join)
            if "status" in df.columns
            else ("order_id", lambda s: "")
        ),
        "symbols": (
            ("symbol", _safe_join)
            if "symbol" in df.columns
            else ("order_id", lambda s: "")
        ),
        "instructions": (
            ("instruction", _safe_join)
            if "instruction" in df.columns
            else ("order_id", lambda s: "")
        ),
        "legs": ("row_kind", lambda s: int((s == "LEG").sum())),
        "containers": ("row_kind", lambda s: int((s == "CONTAINER").sum())),
        "open_rows": (
            "is_open_status",
            lambda s: int(pd.Series(s).fillna(False).astype(bool).sum()),
        ),
        "actionable_legs": (
            "is_actionable_leg",
            lambda s: int(pd.Series(s).fillna(False).astype(bool).sum()),
        ),
        "cancelable_any": (
            ("cancelable", _bool_any)
            if "cancelable" in df.columns
            else ("order_id", lambda s: False)
        ),
        "editable_any": (
            ("editable", _bool_any)
            if "editable" in df.columns
            else ("order_id", lambda s: False)
        ),
    }

    gb = df.groupby(keys, dropna=False).agg(**agg).reset_index()

    if "entered_time_min" in gb.columns:
        gb = gb.sort_values(
            by=["actionable_legs", "open_rows", "entered_time_min"],
            ascending=[False, False, False],
            na_position="last",
        )
    return gb


# ----------------------------
# Snapshot changes (diff)
# ----------------------------

def build_snapshot_changes(df_snap: pd.DataFrame) -> pd.DataFrame:
    OUT_COLS = [
        "account_hash",
        "order_id",
        "leg_index",
        "prev_snapshot_ts",
        "last_snapshot_ts",
        "changed_fields",
        "status_prev",
        "status_last",
        "remaining_qty_prev",
        "remaining_qty_last",
        "price_prev",
        "price_last",
        "stop_price_prev",
        "stop_price_last",
        "symbol_prev",
        "symbol_last",
    ]

    if df_snap is None or df_snap.empty:
        return pd.DataFrame(columns=OUT_COLS)

    d = df_snap.copy()
    need = ["account_hash", "order_id", "leg_index", "snapshot_ts"]
    if any(c not in d.columns for c in need):
        return pd.DataFrame(columns=OUT_COLS)

    d["snapshot_ts"] = pd.to_datetime(d["snapshot_ts"], errors="coerce")
    d = d.sort_values(["account_hash", "order_id", "leg_index", "snapshot_ts"])

    watch = ["status", "remaining_qty", "price", "stop_price", "symbol"]

    out_rows = []
    for (ah, oid, li), g in d.groupby(
        ["account_hash", "order_id", "leg_index"], dropna=False
    ):
        g2 = g.tail(2)
        last = g2.iloc[-1]
        prev = g2.iloc[-2] if len(g2) == 2 else None

        changed: List[str] = []
        if prev is not None:
            for c in watch:
                if c not in g2.columns:
                    continue
                a = prev[c]
                b = last[c]
                if (pd.isna(a) and pd.isna(b)) or (a == b):
                    continue
                changed.append(c)

        out_rows.append(
            {
                "account_hash": ah,
                "order_id": oid,
                "leg_index": li,
                "prev_snapshot_ts": (
                    prev["snapshot_ts"] if prev is not None else pd.NaT
                ),
                "last_snapshot_ts": last["snapshot_ts"],
                "changed_fields": (
                    ", ".join(changed)
                    if changed
                    else ("FIRST_SEEN" if prev is None else "")
                ),
                "status_prev": (prev.get("status") if prev is not None else None),
                "status_last": last.get("status"),
                "remaining_qty_prev": (
                    prev.get("remaining_qty") if prev is not None else None
                ),
                "remaining_qty_last": last.get("remaining_qty"),
                "price_prev": (prev.get("price") if prev is not None else None),
                "price_last": last.get("price"),
                "stop_price_prev": (
                    prev.get("stop_price") if prev is not None else None
                ),
                "stop_price_last": last.get("stop_price"),
                "symbol_prev": (prev.get("symbol") if prev is not None else None),
                "symbol_last": last.get("symbol"),
            }
        )

    out = pd.DataFrame(out_rows, columns=OUT_COLS)
    if out.empty:
        return pd.DataFrame(columns=OUT_COLS)

    return out.sort_values(
        ["last_snapshot_ts", "order_id", "leg_index"], ascending=[False, True, True]
    )


# ----------------------------
# Dashboard (user-first, from RAW)
# ----------------------------

def build_order_dashboard(df_raw_flat: pd.DataFrame) -> pd.DataFrame:
    """
    Dashboard output (user-friendly headers):
      Order ID, Parent Order ID, Root Order ID, Placed Date, Strategy, Complex Strategy,
      Instruction, Quantity, Underlying, Expiry, Strike, Order Type, Put/Call,
      Limit Price, Stop Price, Duration, Mark, Status

    - Expiry/Strike: use expirationDate/strikePrice; if missing, parse from OCC symbol
    - IDs are Int64 (Excel format enforced later to prevent scientific notation)
    """
    if df_raw_flat is None or df_raw_flat.empty:
        return pd.DataFrame(columns=DASHBOARD_COLUMNS)

    df = df_raw_flat.copy()
    if "order_id" not in df.columns:
        return pd.DataFrame(columns=DASHBOARD_COLUMNS)

    # Root grouping key
    df["_root_id"] = (
        df["root_order_id"].fillna(df["order_id"])
        if "root_order_id" in df.columns
        else df["order_id"]
    )

    # Eligible roots: any row in the root with a dashboard status
    if "status" in df.columns:
        eligible_roots = set(
            df.loc[df["status"].isin(DASHBOARD_STATUSES), "_root_id"].dropna().tolist()
        )
    else:
        eligible_roots = set()

    if not eligible_roots:
        return pd.DataFrame(columns=DASHBOARD_COLUMNS)

    # Keep only actionable statuses in those roots
    df = df[df["_root_id"].isin(eligible_roots)].copy()
    df = df[df["status"].isin(DASHBOARD_STATUSES)].copy()

    if df.empty:
        return pd.DataFrame(columns=DASHBOARD_COLUMNS)

    placed = df["enteredTime"] if "enteredTime" in df.columns else None
    placed_date = (
        placed.map(_fmt_yyyy_mm_dd)
        if placed is not None
        else pd.Series([None] * len(df), index=df.index)
    )

    expiry = df.get("expirationDate", pd.Series([None] * len(df), index=df.index))
    strike = df.get("strikePrice", pd.Series([None] * len(df), index=df.index))
    put_call = df.get("putCall", pd.Series([None] * len(df), index=df.index))
    und = df.get("underlyingSymbol", pd.Series([None] * len(df), index=df.index))

    # Fallback parse from symbol if expiry/strike missing
    sym = df.get("symbol", pd.Series([None] * len(df), index=df.index))
    need_fallback = expiry.isna() | strike.isna() | put_call.isna() | und.isna()
    if need_fallback.any():
        parsed = [ _parse_occ_symbol_for_expiry_strike(sym.iloc[i]) for i in range(len(df)) ]
        p_exp = pd.Series([t[0] for t in parsed], index=df.index)
        p_str = pd.Series([t[1] for t in parsed], index=df.index)
        p_pc  = pd.Series([t[2] for t in parsed], index=df.index)
        p_und = pd.Series([t[3] for t in parsed], index=df.index)

        expiry = expiry.where(~expiry.isna(), p_exp)
        strike = strike.where(~strike.isna(), p_str)
        put_call = put_call.where(~put_call.isna(), p_pc)
        und = und.where(~und.isna(), p_und)

    expiry = expiry.map(_fmt_yyyy_mm_dd)
    strike = pd.to_numeric(strike, errors="coerce")

    out = pd.DataFrame(
        {
            "Order ID": df.get("order_id"),
            "Parent Order ID": df.get("parent_order_id"),
            "Root Order ID": df["_root_id"],
            "Placed Date": placed_date,
            "Strategy": df.get("orderStrategyType"),
            "Complex Strategy": df.get("complexOrderStrategyType"),
            "Instruction": df.get("instruction"),
            "Quantity": df.get("quantity"),
            "Underlying": und,
            "Expiry": expiry,
            "Strike": strike,
            "Order Type": df.get("orderType"),
            "Put/Call": put_call,
            "Limit Price": df.get("price"),
            "Stop Price": df.get("stopPrice"),
            "Duration": df.get("duration"),
            "Mark": None,   # wired separately (see notes)
            "Status": df.get("status"),
        }
    )

    # IDs as Int64 (preserve blanks)
    for c in ["Order ID", "Parent Order ID", "Root Order ID"]:
        out[c] = _to_int64_nullable(out[c])

    # Sort inside dashboard to group roots cleanly
    def _strategy_rank(v: Any) -> int:
        s = str(v) if v is not None else ""
        if s == "OCO":
            return 0
        if s == "SINGLE":
            return 1
        return 9

    def _status_rank(v: Any) -> int:
        s = str(v) if v is not None else ""
        if s == "PENDING_ACTIVATION":
            return 0
        if s == "WORKING":
            return 1
        return 9

    out["_sr"] = out["Strategy"].map(_strategy_rank)
    out["_str"] = out["Status"].map(_status_rank)

    sort_cols = ["Placed Date", "Root Order ID", "_sr", "_str", "Order ID"]
    out = out.sort_values(by=sort_cols, ascending=[False, True, True, True, True], na_position="last")
    out = out.drop(columns=["_sr", "_str"], errors="ignore")

    return out.reindex(columns=DASHBOARD_COLUMNS)


# ----------------------------
# Summary
# ----------------------------

def build_summary(
    df_live: pd.DataFrame,
    df_raw_meta: pd.DataFrame,
    df_snap: pd.DataFrame,
    df_raw_flat: pd.DataFrame,
) -> pd.DataFrame:
    distinct_live_orders = (
        int(df_live["order_id"].nunique())
        if (df_live is not None and not df_live.empty and "order_id" in df_live.columns)
        else 0
    )
    distinct_snap_orders = (
        int(df_snap["order_id"].nunique())
        if (df_snap is not None and not df_snap.empty and "order_id" in df_snap.columns)
        else 0
    )

    raw_missing_parent = 0
    if (
        df_raw_flat is not None
        and not df_raw_flat.empty
        and {"depth", "parent_order_id"}.issubset(df_raw_flat.columns)
    ):
        raw_missing_parent = int(
            ((df_raw_flat["depth"] > 0) & (df_raw_flat["parent_order_id"].isna())).sum()
        )

    live_missing_parent = 0
    if (
        df_live is not None
        and not df_live.empty
        and {"depth", "parent_order_id"}.issubset(df_live.columns)
    ):
        live_missing_parent = int(
            ((df_live["depth"] > 0) & (df_live["parent_order_id"].isna())).sum()
        )

    snap_missing_parent = 0
    if (
        df_snap is not None
        and not df_snap.empty
        and {"depth", "parent_order_id"}.issubset(df_snap.columns)
    ):
        snap_missing_parent = int(
            ((df_snap["depth"] > 0) & (df_snap["parent_order_id"].isna())).sum()
        )

    latest_raw = {}
    if df_raw_meta is not None and not df_raw_meta.empty:
        r0 = df_raw_meta.iloc[0].to_dict()
        latest_raw = {
            "orders_raw_latest_fetched_at": r0.get("fetched_at"),
            "orders_raw_latest_sha256": r0.get("payload_sha256"),
            "orders_raw_latest_http_status": r0.get("http_status"),
            "orders_raw_latest_status_filter": r0.get("status_filter"),
            "orders_raw_latest_from_iso": r0.get("from_iso"),
            "orders_raw_latest_to_iso": r0.get("to_iso"),
        }

    summary = {
        "run_utc": _now_utc_str(),
        "open_orders_live_rows": int(len(df_live)) if df_live is not None else 0,
        "open_orders_live_distinct_order_id": distinct_live_orders,
        "open_orders_snapshots_rows_in_report": (
            int(len(df_snap)) if df_snap is not None else 0
        ),
        "open_orders_snapshots_distinct_order_id_in_report": distinct_snap_orders,
        "raw_flat_rows": int(len(df_raw_flat)) if df_raw_flat is not None else 0,
        "raw_flat_distinct_order_id": (
            int(df_raw_flat["order_id"].nunique())
            if (
                df_raw_flat is not None
                and not df_raw_flat.empty
                and "order_id" in df_raw_flat.columns
            )
            else 0
        ),
        "raw_missing_parent": raw_missing_parent,
        "live_missing_parent": live_missing_parent,
        "snap_missing_parent": snap_missing_parent,
        **latest_raw,
        "report_lookback_days": REPORT_LOOKBACK_DAYS,
    }
    return pd.DataFrame([summary])


# ----------------------------
# Excel write safety + formatting
# ----------------------------

def _excel_safe_datetimes(dfx: pd.DataFrame) -> pd.DataFrame:
    """
    Excel can't handle tz-aware datetimes.
    - Converts datetime64[ns, tz] -> naive UTC
    - Also handles object columns containing tz-aware datetime/Timestamp
    """
    if dfx is None or dfx.empty:
        return dfx

    out = dfx.copy()

    for c in out.columns:
        s = out[c]

        if isinstance(s.dtype, pd.DatetimeTZDtype):
            out[c] = s.dt.tz_convert("UTC").dt.tz_localize(None)
            continue

        if s.dtype == "object":
            sample = s.dropna().head(50).tolist()
            if not sample:
                continue

            has_tz = False
            for v in sample:
                if isinstance(v, pd.Timestamp) and v.tzinfo is not None:
                    has_tz = True
                    break
                if isinstance(v, datetime) and v.tzinfo is not None:
                    has_tz = True
                    break

            if not has_tz:
                continue

            def _strip_tz(v: Any) -> Any:
                if v is None:
                    return None
                if isinstance(v, pd.Timestamp):
                    if v.tzinfo is None:
                        return v.to_pydatetime()
                    return v.tz_convert("UTC").tz_localize(None).to_pydatetime()
                if isinstance(v, datetime):
                    if v.tzinfo is None:
                        return v
                    return v.astimezone(timezone.utc).replace(tzinfo=None)
                return v

            out[c] = s.map(_strip_tz)

    return out


def _format_dashboard_sheet(ws) -> None:
    """
    Excel styling:
      - Header pastel blue
      - Alternate root group fill pastel green / none / green ...
      - Force ID number format "0" (prevents scientific notation)
      - Date format yyyy-mm-dd for Placed Date and Expiry
      - Auto-fit columns
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="D9E1F2")  # pastel blue
    group_fill = PatternFill("solid", fgColor="E2EFDA")   # pastel green

    max_row = ws.max_row
    max_col = ws.max_column

    ws.freeze_panes = "A2"

    # Header style
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}
    root_idx = headers.get("Root Order ID")
    id_cols = [
        headers.get(x)
        for x in ["Order ID", "Parent Order ID", "Root Order ID"]
        if headers.get(x) is not None
    ]
    date_cols = [
        headers.get(x) for x in ["Placed Date", "Expiry"] if headers.get(x) is not None
    ]
    mark_col = headers.get("Mark")

    # Alternate group fill based on Root Order ID
    toggle = False
    prev_root = None
    if root_idx is not None:
        for r in range(2, max_row + 1):
            v = ws.cell(row=r, column=root_idx).value
            if v != prev_root:
                toggle = not toggle
                prev_root = v
            if toggle:
                for c in range(1, max_col + 1):
                    ws.cell(row=r, column=c).fill = group_fill

    # Number formats
    for r in range(2, max_row + 1):
        for c in id_cols:
            cell = ws.cell(row=r, column=c)
            cell.number_format = "0"
            cell.alignment = Alignment(horizontal="right")

        for c in date_cols:
            cell = ws.cell(row=r, column=c)
            cell.number_format = "yyyy-mm-dd"
            cell.alignment = Alignment(horizontal="left")

        if mark_col is not None:
            ws.cell(row=r, column=mark_col).number_format = "0.00"

    # Auto-fit columns (with caps)
    def _cell_len(v: Any) -> int:
        if v is None:
            return 0
        return len(str(v))

    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            max_len = max(max_len, _cell_len(ws.cell(row=r, column=c).value))
        width = min(max(10, max_len + 2), 60)
        if c in id_cols:
            width = max(width, 22)  # ensure full 13-digit IDs show cleanly
        ws.column_dimensions[letter].width = width

    ws.auto_filter.ref = ws.dimensions


def write_report(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    cleaned: dict[str, pd.DataFrame] = {}
    for name, df in sheets.items():
        if df is None or df.empty:
            cleaned[name] = pd.DataFrame()
            continue

        dfx = df.copy()

        # Convert dict/list objects to strings for Excel
        for c in dfx.columns:
            if dfx[c].dtype == "object":
                sample = dfx[c].dropna().head(10).tolist()
                if any(isinstance(v, (dict, list)) for v in sample):
                    dfx[c] = dfx[c].map(lambda v: str(v) if v is not None else None)

        dfx = _excel_safe_datetimes(dfx)
        cleaned[name] = dfx

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name, df in cleaned.items():
            df.to_excel(xw, sheet_name=name[:31], index=False)

        if "Dashboard" in xw.sheets:
            _format_dashboard_sheet(xw.sheets["Dashboard"])

    print(f"[report_open_orders_live] Wrote: {path}")


# ----------------------------
# Main
# ----------------------------

def main() -> None:
    con = duckdb.connect(get_db_path())

    # LIVE
    df_live = build_live_derived(read_open_orders_live(con))
    df_grouped = build_grouped_orders(df_live)

    df_open_only = (
        df_live[df_live["is_open_status"]].copy()
        if (not df_live.empty and "is_open_status" in df_live.columns)
        else pd.DataFrame()
    )
    df_actionable = (
        df_live[df_live["is_actionable_leg"]].copy()
        if (not df_live.empty and "is_actionable_leg" in df_live.columns)
        else pd.DataFrame()
    )

    # RAW meta + RAW flatten (latest payload)  ✅ always defines df_raw_flat
    df_raw_meta = read_latest_orders_raw_meta(con)
    raw_meta, raw_payload = read_latest_orders_raw_payload(con)
    df_raw_flat = flatten_latest_raw_payload(raw_meta or {}, raw_payload) if raw_meta else pd.DataFrame()

    # DASHBOARD (from RAW)
    df_dash = build_order_dashboard(df_raw_flat)

    # SNAPSHOTS history + diff
    df_snap = read_open_orders_snapshots(con, REPORT_LOOKBACK_DAYS)
    if not df_snap.empty and "snapshot_ts" in df_snap.columns:
        df_snap["snapshot_ts"] = pd.to_datetime(df_snap["snapshot_ts"], errors="coerce")
    df_changes = build_snapshot_changes(df_snap) if not df_snap.empty else pd.DataFrame()

    # SUMMARY
    df_summary = build_summary(df_live, df_raw_meta, df_snap, df_raw_flat)

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M%S")
    out_path = out_dir() / f"{ts}_open_orders_audit_report.xlsx"

    sheets = {
        "Summary": df_summary,
        "Dashboard": df_dash,
        "Raw": df_raw_flat,
        "Live_Legs_All": df_live,
        "Live_Open_Only": df_open_only,
        "Live_Actionable_Legs": df_actionable,
        "Grouped_Orders": df_grouped,
        "Snapshots_Legs": df_snap,
        "Snapshot_Changes": df_changes,
        "Orders_Raw_Meta": df_raw_meta,
    }

    write_report(out_path, sheets)


if __name__ == "__main__":
    main()
